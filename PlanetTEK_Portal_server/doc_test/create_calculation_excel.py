import copy
import datetime
import json
import os
import re
import openpyxl


def parse_offer_number(offer_number):
    """offer_number string'ini ayrıştırarak module_details1 ve module_details2
    değerlerini üretir.
    """
    parts = offer_number.strip().split()
    if not parts:
        return "", ""

    module_details2 = parts[0]

    date_start_idx = -1
    for i in range(len(parts) - 2):
        if parts[i].isdigit() and parts[i + 1].isdigit() and parts[i + 2].isdigit():
            if len(parts[i + 2]) == 4:
                date_start_idx = i
                break

    module_details1 = ""
    if date_start_idx != -1:
        after_date_parts = parts[date_start_idx + 3 :]

        model_parts = []
        for p in after_date_parts:
            if p.isdigit() and len(p) >= 2 and int(p) > 10:
                break
            model_parts.append(p)

        module_details1 = " ".join(model_parts)

    return module_details1, module_details2


def get_capacity_string(form_data):
    """planetDiskDetails.debi değerini formatlar."""
    customer_info = form_data.get("customerInfo", {})
    planet_disk = form_data.get("planetDiskDetails", {})

    unit_system = str(customer_info.get("unitSystem", "")).strip().upper()
    teklif_dili = str(customer_info.get("teklifDili", "")).strip().lower()

    raw_debi = planet_disk.get("debi", 0)
    try:
        debi_val = float(raw_debi)
    except (ValueError, TypeError):
        debi_val = 0.0

    if unit_system in ["US", "IMPERIAL"]:
        gpd_val = round(debi_val * 264.172)
        return f"{gpd_val:,}".replace(",", ".") + " GPD"
    else:
        formatted_debi = f"{int(debi_val)}" if debi_val.is_integer() else f"{debi_val}"
        unit_label = "m³/gün" if teklif_dili == "yerli" else "m³/day"
        return f"{formatted_debi} {unit_label}"


def get_excel_filename_and_folder(json_path):
    """Hem klasör adını hem de Word belgesiyle birebir uyumlu Excel dosya adını üretir."""
    if not os.path.exists(json_path):
        raise FileNotFoundError(f"'{json_path}' bulunamadı!")

    with open(json_path, "r", encoding="utf-8") as f:
        form_data = json.load(f)

    customer_info = form_data.get("customerInfo", {})
    customer_id = customer_info.get("customer_id")
    company_name = customer_info.get("ticari_unvan", "")
    offer_number = customer_info.get("offer_number", "").strip()
    teklif_dili_raw = customer_info.get("teklifDili", "").strip()
    teklif_dili = teklif_dili_raw.lower()

    # 1. Klasör Adı Yapılandırması: {customer_id} - {offer_number}
    if customer_id and offer_number:
        folder_name_raw = f"{customer_id} - {offer_number}"
    elif offer_number:
        folder_name_raw = str(offer_number)
    else:
        folder_name_raw = "Teklif_Klasoru"

    folder_name = re.sub(r'[\\/*?:"<>|]', "", folder_name_raw)

    # 2. Excel Dosya Adı Yapılandırması (Word dökümanı ile birebir aynı)
    module_details1, module_details2 = parse_offer_number(offer_number)
    compacity = get_capacity_string(form_data)
    date_today_short = datetime.datetime.now().strftime("%d.%m.%Y")

    if teklif_dili == "yerli":
        filename_raw = f"PlanetTEK Teklif {company_name} - {module_details1} - {module_details2} - {date_today_short} - {compacity}.xlsx"
    else:
        filename_raw = f"PlanetTEK Offer {company_name} - {module_details1} - {module_details2} - {date_today_short} - {compacity}.xlsx"

    file_name = re.sub(r'[\\/*?:"<>|]', "", filename_raw)

    return folder_name, file_name


def copy_sheet_with_styles(source_sheet, target_sheet):
    """Hücre değerlerini, renkleri, fontları, kenarlıkları, birleştirilmiş
    hücreleri ve sütun genişliklerini birebir hedef sekmeye kopyalar.
    """
    for row in source_sheet.iter_rows():
        for cell in row:
            target_cell = target_sheet.cell(
                row=cell.row, column=cell.column, value=cell.value
            )

            if cell.has_style:
                target_cell.font = copy.copy(cell.font)
                target_cell.border = copy.copy(cell.border)
                target_cell.fill = copy.copy(cell.fill)
                target_cell.number_format = copy.copy(cell.number_format)
                target_cell.protection = copy.copy(cell.protection)
                target_cell.alignment = copy.copy(cell.alignment)

    for merged_range in source_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merged_range))

    for col_key, col_dim in source_sheet.column_dimensions.items():
        target_sheet.column_dimensions[col_key].width = col_dim.width

    for row_key, row_dim in source_sheet.row_dimensions.items():
        target_sheet.row_dimensions[row_key].height = row_dim.height


def create_calculation_excel():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(base_dir, "formData.json")

    xlsx_source_dir = os.path.join(base_dir, "table_creators", "tables", "xlsx1")

    # 🔍 Klasör ve Dosya Adı Hesaplaması
    offer_number_folder, output_filename = get_excel_filename_and_folder(json_path)
    target_dir = os.path.join(base_dir, "final_offer", offer_number_folder)

    os.makedirs(target_dir, exist_ok=True)

    if not os.path.exists(xlsx_source_dir):
        print(f"Hata: Kaynak Excel klasörü '{xlsx_source_dir}' bulunamadı!")
        return

    excel_files = sorted(
        [
            f
            for f in os.listdir(xlsx_source_dir)
            if f.endswith(".xlsx") or f.endswith(".xls")
        ]
    )

    if not excel_files:
        print(f"Hata: '{xlsx_source_dir}' dizininde hiç Excel dosyası bulunamadı!")
        return

    merged_wb = openpyxl.Workbook()
    merged_wb.remove(merged_wb.active)

    # 📄 Yeni Dinamik Dosya Adı
    output_file_path = os.path.join(target_dir, output_filename)

    try:
        for file_name in excel_files:
            file_path = os.path.join(xlsx_source_dir, file_name)
            sheet_name = os.path.splitext(file_name)[0][:31]

            source_wb = openpyxl.load_workbook(file_path)
            source_sheet = source_wb.active

            target_sheet = merged_wb.create_sheet(title=sheet_name)

            copy_sheet_with_styles(source_sheet, target_sheet)

            print(f"Stil ile eklendi: {file_name} -> Sekme: {sheet_name}")

        merged_wb.save(output_file_path)
        print("\nİşlem başarıyla tamamlandı!")
        print(f"Biçimlendirilmiş Excel Oluşturuldu: {output_file_path}")

    except Exception as e:
        print(f"Excel dosyası birleştirilirken hata oluştu: {e}")


if __name__ == "__main__":
    create_calculation_excel()
