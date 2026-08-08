import json
import os
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ==========================================
# 1. DOSYA & JSON VERİ OKUMA (GÜNCELLENDİ)
# ==========================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# 💾 Kaydedilecek hedef dosya ve klasör (xlsx1)
OUTPUT_DIR = os.path.join(BASE_DIR, "..", "tables", "xlsx1")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "parametre.xlsx")

# Target klasör (xlsx1) yoksa otomatik oluştur
os.makedirs(OUTPUT_DIR, exist_ok=True)

JSON_PATH = os.path.join(BASE_DIR, "..", "..", "formData.json")

if not os.path.exists(JSON_PATH):
    raise FileNotFoundError(f"Hata: {JSON_PATH} dosyası bulunamadı!")

with open(JSON_PATH, "r", encoding="utf-8") as f:
    form_data = json.load(f)

customer_info = form_data.get("customerInfo", {})
parametre_data = form_data.get("tables", {}).get("parametretablosu", [])

if not parametre_data:
    print("Uyarı: 'tables.parametretablosu' altında veri bulunamadı veya liste boş!")

# --- DİL SEÇİMİ VE BİRİM SİSTEMİ DİNAMİKLERİ ---
is_foreign = customer_info.get("teklifDili") == "Yabancı"
unit_system = customer_info.get("unitSystem", "metric").upper()


def fix_unit_system(unit_str):
    """unitSystem == 'US' ise parametre tablosundaki metrik birimleri US eşleniklerine çevirir."""
    if not unit_str or unit_str == "-":
        return "-"

    unit_clean = str(unit_str).strip()

    if unit_system == "US":
        unit_clean = (
            unit_clean.replace("m³/saat", "GPH")
            .replace("m3/hour", "GPH")
            .replace("m³/gün", "GPD")
            .replace("m3/day", "GPD")
            .replace("m³", "gal")
            .replace("m3", "gal")
            .replace("kg", "lbs")
        )
    return unit_clean


def fmt_value(val, decimals=2):
    """Değer sayısal ise teklifDili'ne göre formatlar."""
    if val is None or val == "-":
        return "-"
    try:
        num_val = float(val)
        if num_val.is_integer():
            formatted = f"{int(num_val):,}"
        else:
            formatted = f"{num_val:,.{decimals}f}"

        if not is_foreign:
            formatted = formatted.replace(",", "X").replace(".", ",").replace("X", ".")

        return formatted
    except (ValueError, TypeError):
        return str(val)


HEADERS = [
    "Parameter" if is_foreign else "Parametre",
    "unit" if is_foreign else "Birim",
    "Influent (*)" if is_foreign else "Giriş (*)",
    "Effluent (!)" if is_foreign else "Çıkış (!)",
]

# ==========================================
# 2. EXCEL SIFIRDAN HAZIRLIK
# ==========================================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Parametre"
ws.views.sheetView[0].showGridLines = True

# --- STİL TANIMLARI ---
FILL_HEADER_GREY = PatternFill(
    start_color="C0C0C0", end_color="C0C0C0", fill_type="solid"
)
FILL_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

font_header = Font(name="Calibri", size=9, bold=True)
font_value = Font(name="Calibri", size=9, bold=False)

# Kenarlık Stilleri
dashed_black_side = Side(border_style="dashed", color="000000")
thin_black_side = Side(border_style="thin", color="000000")
double_black_side = Side(border_style="double", color="000000")

# ==========================================
# 3. KOLON BAŞLIKLARI (SATIR 1)
# ==========================================
ws.row_dimensions[1].height = 22

for col_idx, header_text in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=col_idx, value=header_text)
    cell.fill = FILL_HEADER_GREY
    cell.font = font_header
    cell.alignment = Alignment(
        horizontal="left" if col_idx == 1 else "center", vertical="center"
    )

    t_top = double_black_side
    t_left = double_black_side if col_idx == 1 else None
    t_right = double_black_side if col_idx == 4 else dashed_black_side
    t_bottom = thin_black_side

    cell.border = Border(top=t_top, bottom=t_bottom, left=t_left, right=t_right)

# ==========================================
# 4. VERİ SATIRLARI DÖNGÜSÜ
# ==========================================
current_row = 2
max_row = len(parametre_data) + 1

for item in parametre_data:
    ws.row_dimensions[current_row].height = 19
    is_last_row = current_row == max_row

    label_val = item.get("label", "")
    unit_raw = item.get("unit", "-") or "-"
    giris_raw = item.get("giriş", "-") or "-"
    cikiss_raw = item.get("çıkış", "-") or "-"

    unit_val = fix_unit_system(unit_raw)
    giris_val = fmt_value(giris_raw)
    cikis_val = fmt_value(cikiss_raw)

    c1 = ws.cell(row=current_row, column=1, value=label_val)
    c2 = ws.cell(row=current_row, column=2, value=unit_val)
    c3 = ws.cell(row=current_row, column=3, value=giris_val)
    c4 = ws.cell(row=current_row, column=4, value=cikis_val)

    c1.alignment = Alignment(horizontal="left", vertical="center")
    c2.alignment = Alignment(horizontal="center", vertical="center")
    c3.alignment = Alignment(horizontal="center", vertical="center")
    c4.alignment = Alignment(horizontal="center", vertical="center")

    for col_idx in range(1, 5):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.fill = FILL_WHITE
        cell.font = font_value

        b_left = double_black_side if col_idx == 1 else None
        b_right = double_black_side if col_idx == 4 else dashed_black_side
        b_bottom = double_black_side if is_last_row else thin_black_side

        cell.border = Border(left=b_left, right=b_right, bottom=b_bottom)

    current_row += 1

# ==========================================
# 5. SÜTUN GENİŞLİKLERİ VE KAYDETME (xlsx1'e)
# ==========================================
ws.column_dimensions["A"].width = 45
ws.column_dimensions["B"].width = 15
ws.column_dimensions["C"].width = 20
ws.column_dimensions["D"].width = 20

wb.save(OUTPUT_FILE)
print(f"🚀 Excel Parametre Tablosu Sıfırdan Çizilerek Kaydedildi: {OUTPUT_FILE}")
