import copy
import json
import os
import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, Side

# ==========================================
# 1. DOSYA & JSON VERİ OKUMA
# ==========================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Orijinal okunacak dosya (xlsx0 altında)
SOURCE_FILE = os.path.join(BASE_DIR, "..", "tables", "xlsx0", "ozet_tablosu.xlsx")
TEMPLATE_FILE = os.path.join(
    BASE_DIR, "..", "tables", "xlsx0", "ozet_tablosu - Kopya.xlsx"
)

# Yeni kaydedilecek hedef klasör ve dosya (xlsx1 altında)
OUTPUT_DIR = os.path.join(BASE_DIR, "..", "tables", "xlsx1")
TARGET_FILE = os.path.join(OUTPUT_DIR, "ozet_tablosu.xlsx")

# Eğer xlsx1 klasörü yoksa otomatik oluştur
os.makedirs(OUTPUT_DIR, exist_ok=True)

JSON_PATH = os.path.join(BASE_DIR, "..", "..", "formData.json")

if not os.path.exists(JSON_PATH):
    raise FileNotFoundError(f"Hata: {JSON_PATH} dosyası bulunamadı!")

with open(JSON_PATH, "r", encoding="utf-8") as f:
    form_data = json.load(f)

customer_info = form_data.get("customerInfo", {})
ozet_data = form_data.get("tables", {}).get("ozettablosu", {})

if not ozet_data:
    ozet_data = form_data.get("tables", {}).get("ozet_tablosu", {})

# ==========================================
# 🎯 DİL KONTROLÜ VE BAŞLIKLAR
# ==========================================
raw_teklif_dili = (
    customer_info.get("teklifDili") or ozet_data.get("teklifDili") or "Yerli"
)
is_turkish = str(raw_teklif_dili).strip().lower() == "yerli"

gen_info = ozet_data.get("generalInfo", {})
offer_no = gen_info.get("offerNo", "")
ref_no = gen_info.get("refNo", "")
client_name = gen_info.get("clientName", customer_info.get("ticari_unvan", ""))

params_list = ozet_data.get("params", [])
content_list = ozet_data.get("content", [])

lbl_offer_no = "Teklif No" if is_turkish else "Quotation No"
lbl_ref_no = "Teklif Referans No" if is_turkish else "Quotation Reference No"
lbl_client = "Müşteri" if is_turkish else "Client"
lbl_design_params = "Tasarım Parametreleri" if is_turkish else "Design Parameters"
lbl_offer_content = "Teklif İçeriği" if is_turkish else "Offer content"

# ==========================================
# 2. TEMİZ ŞABLONU YÜKLEME VE KORUMALAR
# ==========================================
# xlsx0 klasöründeki dosya/şablon yükleniyor
load_path = TEMPLATE_FILE if os.path.exists(TEMPLATE_FILE) else SOURCE_FILE

if not os.path.exists(load_path):
    raise FileNotFoundError(f"Hata: Şablon dosyası bulunamadı! ({load_path})")

wb = openpyxl.load_workbook(load_path)

# Gizli External Link XML başvurusunu temizle (Onarılan kayıtlar uyarısını önler)
if hasattr(wb, "_external_links"):
    wb._external_links.clear()

if "OzetTablosu" in wb.sheetnames:
    ws = wb["OzetTablosu"]
elif "ÖzetTablosu" in wb.sheetnames:
    ws = wb["ÖzetTablosu"]
else:
    ws = wb.active

# Gizli Koşullu Biçimlendirmeleri Temizle
if hasattr(ws, "conditional_formatting") and hasattr(
    ws.conditional_formatting, "_cf_rules"
):
    ws.conditional_formatting._cf_rules.clear()

ws.views.sheetView[0].showGridLines = True


def safe_set(sheet, cell_ref, value):
    cell = sheet[cell_ref]
    if not isinstance(cell, MergedCell):
        cell.value = value


def parse_numeric(val):
    if val is None or val == "" or val == "-":
        return val
    try:
        val_str = str(val).strip()
        if "." in val_str or "," in val_str:
            return float(val_str.replace(",", "."))
        return int(val_str)
    except (ValueError, TypeError):
        return val


# Stil ve Kenarlık Tanımları
EXCEL_WINGDINGS_CHECK = "ü"
wingdings_font = Font(name="Wingdings", size=9.0, bold=False, italic=False)
calibri_normal_font = Font(name="Calibri", size=9.0, bold=False, italic=False)

align_right = Alignment(horizontal="right", vertical="center")
align_center = Alignment(horizontal="center", vertical="center")
align_left = Alignment(horizontal="left", vertical="center")

side_medium = Side(style="medium", color="FF000000")
side_double = Side(style="double", color="FF000000")

# ==========================================
# 3. VERİLERİ ŞABLONA DİNAMİK YAZMA
# ==========================================

# 1. Üst Genel Bilgiler (1 - 3. Satırlar)
safe_set(ws, "A1", lbl_offer_no)
safe_set(ws, "E1", parse_numeric(offer_no))

safe_set(ws, "A2", lbl_ref_no)
safe_set(ws, "E2", ref_no)

safe_set(ws, "A3", lbl_client)
safe_set(ws, "E3", client_name)

# 2. Design Parameters Başlığı (4. Satır)
safe_set(ws, "A4", lbl_design_params)

# 3. Design Parameters Listesi
default_params_count = 17
target_params_count = len(params_list)

if target_params_count < default_params_count:
    ws.delete_rows(5 + target_params_count, default_params_count - target_params_count)
elif target_params_count > default_params_count:
    for _ in range(target_params_count - default_params_count):
        ws.insert_rows(22)

# Params Verilerini Yazma
p_row = 5
for item in params_list:
    label = str(item.get("label", "")).lstrip("-").strip()
    value = parse_numeric(item.get("value", ""))
    unit = item.get("unit", "")

    safe_set(ws, f"A{p_row}", "·")
    safe_set(ws, f"B{p_row}", label)
    safe_set(ws, f"E{p_row}", value)
    safe_set(ws, f"F{p_row}", unit if unit else None)

    for col_letter in ["A", "B", "C", "D", "E", "F"]:
        cell = ws[f"{col_letter}{p_row}"]
        b_curr = cell.border
        top_b = copy.copy(b_curr.top) if b_curr else None
        left_b = copy.copy(b_curr.left) if b_curr else None
        right_b = copy.copy(b_curr.right) if b_curr else None
        cell.border = Border(left=left_b, right=right_b, top=top_b, bottom=side_medium)

    p_row += 1

# 4. Offer Content Başlığı
content_header_row = p_row
safe_set(ws, f"A{content_header_row}", lbl_offer_content)

# 5. Offer Content Listesi
default_content_count = 22
target_content_count = len(content_list)
content_start_row = content_header_row + 1

if target_content_count < default_content_count:
    ws.delete_rows(
        content_start_row + target_content_count,
        default_content_count - target_content_count,
    )
elif target_content_count > default_content_count:
    insert_pos = content_start_row + default_content_count
    for _ in range(target_content_count - default_content_count):
        ws.insert_rows(insert_pos)

# Content Verilerini Yazma
c_row = content_start_row
for item in content_list:
    # 🌟 B VE C HÜCRELERİ BİRLEŞİKSE ÇÖZ (Unmerge):
    merged_ranges_to_remove = []
    for rng in ws.merged_cells.ranges:
        if f"B{c_row}" in str(rng) and f"C{c_row}" in str(rng):
            merged_ranges_to_remove.append(rng)
    for rng in merged_ranges_to_remove:
        ws.unmerge_cells(str(rng))

    # D, E, F sütunlarını birleştir
    try:
        ws.merge_cells(start_row=c_row, start_column=4, end_row=c_row, end_column=6)
    except Exception:
        pass

    is_checked = item.get("isChecked", True)
    is_header = item.get("isHeaderStyle", False)

    qty = parse_numeric(item.get("qty", ""))
    unit = item.get("unit", "")
    desc = item.get("desc", "")

    cell_a = ws[f"A{c_row}"]
    cell_b = ws[f"B{c_row}"]
    cell_c = ws[f"C{c_row}"]
    cell_d = ws[f"D{c_row}"]

    safe_set(ws, f"A{c_row}", None)
    safe_set(ws, f"B{c_row}", None)
    safe_set(ws, f"C{c_row}", None)
    safe_set(ws, f"D{c_row}", None)

    if isinstance(cell_b, openpyxl.cell.cell.Cell):
        cell_b.font = calibri_normal_font
        cell_b.alignment = align_center
    if isinstance(cell_c, openpyxl.cell.cell.Cell):
        cell_c.font = calibri_normal_font
        cell_c.alignment = align_center
    if isinstance(cell_d, openpyxl.cell.cell.Cell):
        cell_d.font = calibri_normal_font
        cell_d.alignment = align_left

    if is_header:
        safe_set(ws, f"D{c_row}", desc)
    else:
        if is_checked and isinstance(cell_a, openpyxl.cell.cell.Cell):
            cell_a.value = EXCEL_WINGDINGS_CHECK
            cell_a.font = wingdings_font
            cell_a.alignment = align_right

        safe_set(ws, f"B{c_row}", qty if qty != "" else "-")
        if unit:
            safe_set(ws, f"C{c_row}", unit)
        safe_set(ws, f"D{c_row}", desc)

    # Ara satır alt kenarlıklarını düzenle
    for col_letter in ["A", "B", "C", "D", "E", "F"]:
        cell = ws[f"{col_letter}{c_row}"]
        b_curr = cell.border
        top_b = copy.copy(b_curr.top) if b_curr else None
        left_b = copy.copy(b_curr.left) if b_curr else None
        right_b = copy.copy(b_curr.right) if b_curr else None
        cell.border = Border(left=left_b, right=right_b, top=top_b, bottom=side_medium)

    c_row += 1

# ==========================================
# 4. TABLONUN EN DIŞ SINIRLARINI ÇİFT ÇİZGİ YAPMA
# ==========================================
final_row = c_row - 1

# Üst Çizgi (1. Satır)
for col_letter in ["A", "B", "C", "D", "E", "F"]:
    cell = ws[f"{col_letter}1"]
    b_curr = cell.border
    cell.border = Border(
        left=copy.copy(b_curr.left) if b_curr else None,
        right=copy.copy(b_curr.right) if b_curr else None,
        top=side_double,
        bottom=copy.copy(b_curr.bottom) if b_curr else None,
    )

# Alt Çizgi (final_row)
for col_letter in ["A", "B", "C", "D", "E", "F"]:
    cell = ws[f"{col_letter}{final_row}"]
    b_curr = cell.border
    cell.border = Border(
        left=copy.copy(b_curr.left) if b_curr else None,
        right=copy.copy(b_curr.right) if b_curr else None,
        top=copy.copy(b_curr.top) if b_curr else None,
        bottom=side_double,
    )

# Sol Çizgi (A Kolonu)
for r in range(1, final_row + 1):
    cell = ws[f"A{r}"]
    b_curr = cell.border
    cell.border = Border(
        left=side_double,
        right=copy.copy(b_curr.right) if b_curr else None,
        top=copy.copy(b_curr.top) if b_curr else None,
        bottom=copy.copy(b_curr.bottom) if b_curr else None,
    )

# Sağ Çizgi (F Kolonu)
for r in range(1, final_row + 1):
    cell = ws[f"F{r}"]
    b_curr = cell.border
    cell.border = Border(
        left=copy.copy(b_curr.left) if b_curr else None,
        right=side_double,
        top=copy.copy(b_curr.top) if b_curr else None,
        bottom=copy.copy(b_curr.bottom) if b_curr else None,
    )

# 🌟 5. SÜTUN GENİŞLİKLERİ AYARI (E genişletildi, F daraltıldı)
ws.column_dimensions["E"].width = 22
ws.column_dimensions["F"].width = 12

# ==========================================
# 6. KAYDETME (xlsx1 KLASÖRÜNE KAYDET)
# ==========================================
wb.save(TARGET_FILE)
print(
    f"Excel Özet Tablosu orijinal dosya bozulmadan yeni yola kaydedildi: {TARGET_FILE}"
)
