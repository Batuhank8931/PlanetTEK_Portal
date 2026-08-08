import json
import os
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ==========================================
# 1. DOSYA & JSON VERİ OKUMA (GÜNCELLENDİ)
# ==========================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# 💾 Kaydedilecek hedef dosya ve klasör (xlsx1)
OUTPUT_DIR = os.path.join(BASE_DIR, "..", "tables", "xlsx1")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "capex.xlsx")

# Target klasör (xlsx1) yoksa otomatik oluştur
os.makedirs(OUTPUT_DIR, exist_ok=True)

JSON_PATH = os.path.join(BASE_DIR, "..", "..", "formData.json")

if not os.path.exists(JSON_PATH):
    raise FileNotFoundError(f"Hata: {JSON_PATH} dosyası bulunamadı!")

with open(JSON_PATH, "r", encoding="utf-8") as f:
    form_data = json.load(f)

data = form_data.get("tables", {}).get("capextablosu", {}).get("rows", [])
customer_info = form_data.get("customerInfo", {})

if not data:
    print("Uyarı: 'tables.capextablosu.rows' altında veri bulunamadı veya liste boş!")

# --- DİL SEÇİMİ VE PARA BİRİMİ DİNAMİKLERİ ---
is_foreign = customer_info.get("teklifDili") == "Yabancı"
currency_code = customer_info.get("currency", "EUR").upper()

if currency_code == "USD":
    CURRENCY_SYMBOL = "$"
elif currency_code == "TRY":
    CURRENCY_SYMBOL = "₺"
else:
    CURRENCY_SYMBOL = "€"


def fmt_num(val, decimals=2):
    try:
        val = float(val)
    except (ValueError, TypeError):
        val = 0.0

    formatted = f"{val:,.{decimals}f}"

    if not is_foreign:
        formatted = formatted.replace(",", "X").replace(".", ",").replace("X", ".")

    return formatted


def format_currency(val):
    if val is None or isinstance(val, str):
        return val or ""
    try:
        val_float = float(val)
        formatted = fmt_num(val_float, 2)
        if currency_code == "USD":
            return f"$ {formatted}"
        elif currency_code == "TRY":
            return f"{formatted} ₺"
        else:
            return f"{formatted} €"
    except (ValueError, TypeError):
        return str(val)


def format_percent(val):
    if val is None:
        return "%0"
    try:
        v = float(val)
        pct = v if v > 1 else v * 100
        if pct.is_integer():
            return f"%{int(pct)}"
        else:
            return f"%{fmt_num(pct, 1)}"
    except (ValueError, TypeError):
        return str(val)


def format_piece(val):
    try:
        val_num = float(val)
        if val_num.is_integer():
            return fmt_num(val_num, 0)
        return fmt_num(val_num, 1)
    except (ValueError, TypeError):
        return str(val or "0")


LANG = {
    "offer_no": "Offer Number" if is_foreign else "Teklif Numarası",
    "offer_ref": (
        "Offer Reference Number" if is_foreign else "Teklif Referans Numarası"
    ),
    "headers": [
        "No",
        "Description" if is_foreign else "Açıklama",
        "Piece" if is_foreign else "Adet",
        "Unit Price" if is_foreign else "Birim Fiyat",
        "Total Price" if is_foreign else "Toplam Fiyat",
        "Discount\nrate" if is_foreign else "İndirim\nOranı",
        ("Total Price\nafter Discount" if is_foreign else "İndirimli\nToplam Fiyat"),
    ],
    "total": "TOTAL" if is_foreign else "TOPLAM",
    "total_discount": "Total Discount" if is_foreign else "Toplam İndirim",
    "grand_total": (
        "Grand Total Discounted Price" if is_foreign else "Genel İndirimli Toplam Fiyat"
    ),
    "notes": [
        (
            "*Optional and Supply Locally equipment are not included in the total price."
            if is_foreign
            else "*Opsiyonel ve Yerel Temin Ekipmanlar toplam fiyata dahil değildir."
        ),
        (
            "* V.A.T. , All Other Expenses and Taxes are EXCLUDED"
            if is_foreign
            else "* K.D.V., Tüm Diğer Masraflar ve Vergiler HARİÇTİR."
        ),
        (
            "* EXWORKS PlanetTEK Kocaeli Gebze Factory"
            if is_foreign
            else "* EXWORKS PlanetTEK Kocaeli Gebze Fabrika Teslimi"
        ),
        ("* Validity : 30 days" if is_foreign else "* Opsiyon Süresi : 30 Gün"),
    ],
    "supply_locally_footer": (
        "Supply locally : equipment should be obtained locally for better warranty conditions."
        if is_foreign
        else "Yerel Temin : Daha iyi garanti koşulları için ekipmanlar yerel olarak temin edilmelidir."
    ),
}

# ==========================================
# 2. EXCEL SIFIRDAN HAZIRLIK
# ==========================================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Capex"
ws.views.sheetView[0].showGridLines = True

# Stil Tanımları
FILL_HEADER_YELLOW = PatternFill(
    start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"
)
FILL_LEVEL_0 = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")
FILL_LEVEL_1 = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
FILL_LEVEL_2 = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
FILL_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

FONT_BASE = {"name": "Calibri", "size": 8}
font_level_0 = Font(bold=True, italic=True, **FONT_BASE)
font_level_1 = Font(bold=True, **FONT_BASE)
font_level_2 = Font(bold=False, italic=True, **FONT_BASE)
font_item = Font(bold=False, **FONT_BASE)
font_item_bold_italic = Font(bold=True, italic=True, **FONT_BASE)
font_header_title = Font(name="Calibri", size=9, bold=True, italic=True)
font_summary_bold = Font(name="Calibri", size=9, bold=True)
font_footer_text = Font(name="Calibri", size=9, bold=False)

thin_side = Side(border_style="thin", color="000000")
border_thin_all = Border(
    left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
)
border_empty = Border()

# ==========================================
# 3. ÜST BİLGİ ALANI (Satır 1, 2, 3)
# ==========================================
ws.row_dimensions[1].height = 18
ws.row_dimensions[2].height = 18
ws.row_dimensions[3].height = 18

for r in range(1, 4):
    for c in range(1, 8):
        cell = ws.cell(row=r, column=c)
        cell.fill = FILL_HEADER_YELLOW
        cell.font = font_header_title
        cell.border = Border(
            top=thin_side if r == 1 else None,
            bottom=thin_side if r == 3 else None,
            left=thin_side if c == 1 else None,
            right=thin_side if c == 7 else None,
        )

ws.merge_cells("A1:G1")
ws["A1"] = "CAPEX"
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells("A2:E2")
ws["A2"] = customer_info.get("ticari_unvan", "")
ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells("F2:G2")
ws["F2"] = f"{LANG['offer_no']} : {customer_info.get('teklifNo', '')}"
ws["F2"].alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells("A3:G3")
ws["A3"] = f"{LANG['offer_ref']} : {customer_info.get('offer_number', '')}"
ws["A3"].alignment = Alignment(horizontal="center", vertical="center")

# ==========================================
# 4. KOLON BAŞLIKLARI (Satır 4)
# ==========================================
ws.row_dimensions[4].height = 28
for col_idx, header_text in enumerate(LANG["headers"], 1):
    cell = ws.cell(row=4, column=col_idx, value=header_text)
    cell.font = Font(name="Calibri", size=8, bold=True)
    cell.fill = FILL_HEADER_YELLOW
    cell.border = border_thin_all
    cell.alignment = Alignment(
        horizontal="left" if col_idx == 2 else "center",
        vertical="center",
        wrap_text=True,
    )

# ==========================================
# 5. VERİ DÖNGÜSÜ
# ==========================================
current_row = 5

total_price_sum = 0.0
grand_total_discounted_sum = 0.0

for item in data:
    is_header = (item.get("piece") == 0 and item.get("unitPrice") == 0) or item.get(
        "type"
    ) in [0, 1, 2]
    label_value = item.get("label", "")

    if item.get("type") == 0:
        label_value = label_value.upper()

    ws.cell(row=current_row, column=1, value=str(item.get("computedNo", "")))
    ws.cell(row=current_row, column=2, value=str(label_value))

    line_count = str(label_value).count("\n") + 1
    ws.row_dimensions[current_row].height = max(18, line_count * 12.5)

    if is_header:
        if item.get("type") == 0 or "." not in item.get("computedNo", "").strip("."):
            row_fill, row_font = FILL_LEVEL_0, font_level_0
        elif item.get("computedNo", "").count(".") == 2:
            row_fill, row_font = FILL_LEVEL_1, font_level_1
        else:
            row_fill, row_font = FILL_LEVEL_2, font_level_2

        for col in range(1, 8):
            cell = ws.cell(row=current_row, column=col)
            cell.fill = row_fill
            cell.font = row_font
            cell.border = border_empty
            cell.alignment = Alignment(
                horizontal="center" if col == 1 else "left",
                vertical="center",
                wrap_text=True,
            )
    else:
        if item.get("isUrgent"):
            for col in range(1, 8):
                cell = ws.cell(row=current_row, column=col)
                cell.fill = FILL_WHITE
                cell.font = font_item
                cell.border = border_empty
                cell.alignment = Alignment(
                    horizontal="center" if col == 1 else "left",
                    vertical="center",
                    wrap_text=True,
                )
        else:
            piece = float(item.get("piece", 0) or 0)
            unit_price = float(item.get("unitPrice", 0.0) or 0.0)
            raw_discount = (
                item.get("discount")
                if item.get("discount") is not None
                else item.get("discountRate", 0)
            )

            try:
                discount_val = float(raw_discount)
            except (ValueError, TypeError):
                discount_val = 0.0

            discount_rate = discount_val / 100.0 if discount_val > 1 else discount_val

            is_special_text = item.get("isOptional") or item.get("isLocalSupply")
            text_value = str(item.get("netTotal", "")) if is_special_text else ""

            ws.cell(row=current_row, column=3, value=format_piece(piece))
            ws.cell(row=current_row, column=4, value=format_currency(unit_price))
            ws.cell(row=current_row, column=6, value=format_percent(discount_val))

            if is_special_text:
                ws.cell(row=current_row, column=5, value=text_value)
                ws.cell(row=current_row, column=7, value=text_value)
            else:
                row_total = piece * unit_price
                row_discounted = row_total * (1.0 - discount_rate)

                total_price_sum += row_total
                grand_total_discounted_sum += row_discounted

                ws.cell(
                    row=current_row,
                    column=5,
                    value=format_currency(row_total),
                )
                ws.cell(
                    row=current_row,
                    column=7,
                    value=format_currency(row_discounted),
                )

            for col in range(1, 8):
                cell = ws.cell(row=current_row, column=col)
                cell.fill = FILL_WHITE
                cell.border = border_empty
                cell.font = (
                    font_item_bold_italic
                    if (col in [5, 7] and is_special_text)
                    else font_item
                )
                cell.alignment = Alignment(
                    horizontal="left" if col == 2 else "center",
                    vertical="center",
                    wrap_text=True,
                )

    current_row += 1

# ==========================================
# 6. URGENT GRUPLARI BİRLEŞTİRME
# ==========================================
search_row = 5
urgent_groups = []
current_group = []

for item in data:
    if item.get("isUrgent"):
        current_group.append((search_row, item))
    else:
        if current_group:
            urgent_groups.append(current_group)
            current_group = []
    search_row += 1
if current_group:
    urgent_groups.append(current_group)

for group in urgent_groups:
    start_r, end_r = group[0][0], group[-1][0]
    unit_data_text = str(group[0][1].get("unitData", ""))
    ws.merge_cells(start_row=start_r, start_column=3, end_row=end_r, end_column=7)
    merged_cell = ws.cell(row=start_r, column=3, value=unit_data_text)
    merged_cell.alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    merged_cell.font = font_item

# ==========================================
# 7. ÖZET VE ALT BİLGİ SATIRLARI
# ==========================================
total_row = current_row
ws.row_dimensions[total_row].height = 20

ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
total_label_cell = ws.cell(row=total_row, column=1, value=LANG["total"])
total_label_cell.alignment = Alignment(horizontal="right", vertical="center")
total_label_cell.font = font_summary_bold

cell_total_e = ws.cell(row=total_row, column=5, value=format_currency(total_price_sum))
cell_total_e.alignment = Alignment(horizontal="center", vertical="center")
cell_total_e.font = font_summary_bold

ws.cell(row=total_row, column=6, value=LANG["total_discount"]).alignment = Alignment(
    horizontal="center", vertical="center"
)
ws.cell(row=total_row, column=6).font = font_summary_bold

current_row += 1
footer_start_row = current_row

for idx, note in enumerate(LANG["notes"]):
    r_idx = footer_start_row + idx
    ws.row_dimensions[r_idx].height = 18
    ws.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=4)
    cell = ws.cell(row=r_idx, column=1, value=note)
    cell.font = font_footer_text
    cell.alignment = Alignment(horizontal="left", vertical="center")

ws.merge_cells(
    start_row=footer_start_row,
    start_column=5,
    end_row=footer_start_row + 3,
    end_column=6,
)
gt_label_cell = ws.cell(row=footer_start_row, column=5, value=LANG["grand_total"])
gt_label_cell.alignment = Alignment(
    horizontal="center", vertical="center", wrap_text=True
)
gt_label_cell.font = font_summary_bold

ws.merge_cells(
    start_row=footer_start_row,
    start_column=7,
    end_row=footer_start_row + 3,
    end_column=7,
)

gt_val_cell = ws.cell(
    row=footer_start_row,
    column=7,
    value=format_currency(grand_total_discounted_sum),
)
gt_val_cell.alignment = Alignment(horizontal="center", vertical="center")
gt_val_cell.font = font_summary_bold

total_discount_amount = total_price_sum - grand_total_discounted_sum
cell_total_g = ws.cell(
    row=total_row, column=7, value=format_currency(total_discount_amount)
)
cell_total_g.alignment = Alignment(horizontal="center", vertical="center")
cell_total_g.font = font_summary_bold

for col in range(1, 8):
    c = ws.cell(row=total_row, column=col)
    c.fill = FILL_HEADER_YELLOW
    c.border = border_thin_all

for r in range(footer_start_row, footer_start_row + 4):
    for c in range(5, 8):
        cell = ws.cell(row=r, column=c)
        cell.fill = FILL_HEADER_YELLOW
        cell.border = border_thin_all

current_row = footer_start_row + 4

ws.row_dimensions[current_row].height = 20
ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
sl_cell = ws.cell(row=current_row, column=1, value=LANG["supply_locally_footer"])
sl_cell.alignment = Alignment(horizontal="center", vertical="center")
sl_cell.font = font_footer_text

current_row += 1

# ==========================================
# 8. EN DIŞ ÇERÇEVE
# ==========================================
max_row = current_row - 1
max_col = 7

for r in range(1, max_row + 1):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=r, column=c)
        top = cell.border.top.style if cell.border and cell.border.top else None
        bottom = (
            cell.border.bottom.style if cell.border and cell.border.bottom else None
        )
        left = cell.border.left.style if cell.border and cell.border.left else None
        right = cell.border.right.style if cell.border and cell.border.right else None

        if r == 1:
            top = "double"
        if r == max_row:
            bottom = "double"
        if c == 1:
            left = "double"
        if c == max_col:
            right = "double"

        cell.border = Border(
            top=Side(border_style=top, color="000000") if top else None,
            bottom=(Side(border_style=bottom, color="000000") if bottom else None),
            left=Side(border_style=left, color="000000") if left else None,
            right=Side(border_style=right, color="000000") if right else None,
        )

# ==========================================
# 9. SÜTUN GENİŞLİKLERİ VE KAYDETME (xlsx1'e)
# ==========================================
ws.column_dimensions["A"].width = 7
ws.column_dimensions["B"].width = 40
ws.column_dimensions["C"].width = 6
ws.column_dimensions["D"].width = 10
ws.column_dimensions["E"].width = 10
ws.column_dimensions["F"].width = 10
ws.column_dimensions["G"].width = 10

wb.save(OUTPUT_FILE)
print(f"🚀 Excel CAPEX Tablosu Sıfırdan Çizilerek Kaydedildi: {OUTPUT_FILE}")
