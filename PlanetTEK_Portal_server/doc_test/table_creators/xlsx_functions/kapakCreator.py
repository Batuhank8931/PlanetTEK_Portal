import json
import os
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ==========================================
# 1. DOSYA & JSON VERİ OKUMA (GÜNCELLENDİ)
# ==========================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# 💾 Kaydedilecek hedef dosya ve klasör (xlsx1)
OUTPUT_DIR = os.path.join(BASE_DIR, "..", "tables", "xlsx1")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "kapak.xlsx")

# Target klasör (xlsx1) yoksa otomatik oluştur
os.makedirs(OUTPUT_DIR, exist_ok=True)

JSON_PATH = os.path.join(BASE_DIR, "..", "..", "formData.json")

if not os.path.exists(JSON_PATH):
    raise FileNotFoundError(f"Hata: {JSON_PATH} dosyası bulunamadı!")

with open(JSON_PATH, "r", encoding="utf-8") as f:
    form_data = json.load(f)

customer_info = form_data.get("customerInfo", {})
kapak_data = form_data.get("tables", {}).get("kapaktablosu", [])

# --- DİL SEÇİMİ, PARA BİRİMİ VE BİRİM SİSTEMİ DİNAMİKLERİ ---
is_foreign = customer_info.get("teklifDili") == "Yabancı"
currency_code = customer_info.get("currency", "EUR").upper()
unit_system = customer_info.get("unitSystem", "metric").upper()

if currency_code == "USD":
    CURRENCY_SYMBOL = "$"
elif currency_code == "TRY":
    CURRENCY_SYMBOL = "₺"
else:
    CURRENCY_SYMBOL = "€"


def fmt_num(val, decimals=2):
    """Sayıyı teklifDili ve 'decimals' hanesine göre formatlar."""
    try:
        val = float(val)
    except (ValueError, TypeError):
        val = 0.0

    formatted = f"{val:,.{decimals}f}"

    if not is_foreign:
        formatted = formatted.replace(",", "X").replace(".", ",").replace("X", ".")

    return formatted


def fix_unit_text(unit_str):
    """unit_str içindeki para birimini ve unitSystem değerine göre birimleri günceller."""
    if not unit_str:
        return ""

    unit_clean = str(unit_str).replace("€", CURRENCY_SYMBOL).strip()

    if unit_system == "US":
        unit_clean = (
            unit_clean.replace("m³/saat", "GPH")
            .replace("m3/hour", "GPH")
            .replace("m³/gün", "GPD")
            .replace("m3/day", "GPD")
            .replace("m³", "gal")
            .replace("m3", "gal")
            .replace("kg", "lbs")
            .replace("ton", "tons")
        )

    return unit_clean


header_title = "OFFER INFORMATION" if is_foreign else "TEKLİF BİLGİLERİ"
offer_no_label = "Offer Number" if is_foreign else "Teklif Numarası"
offer_no_value = customer_info.get("offer_number", "")

# ==========================================
# 2. EXCEL SIFIRDAN HAZIRLIK
# ==========================================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Kapak"
ws.views.sheetView[0].showGridLines = True

# --- STİL TANIMLARI ---
FILL_HEADER_YELLOW = PatternFill(
    start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"
)
FILL_LABEL_GREY = PatternFill(
    start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
)
FILL_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

font_header = Font(name="Calibri", size=10, bold=True, italic=True)
font_label = Font(name="Calibri", size=9, bold=True, italic=True)
font_value = Font(name="Calibri", size=9, bold=False)

thin_grey_side = Side(border_style="thin", color="E0E0E0")
double_black_side = Side(border_style="double", color="000000")

# ==========================================
# 3. VERİ LİSTESİ HAZIRLIĞI
# ==========================================
rows_to_render = []

if offer_no_value:
    rows_to_render.append({"label": offer_no_label, "value_str": str(offer_no_value)})

for item in kapak_data:
    lbl = item.get("label", "")
    val = item.get("value", "")
    unit_raw = item.get("unit", "")

    unit_fixed = fix_unit_text(unit_raw)

    try:
        val_num = float(val)
        val_str_formatted = (
            fmt_num(val_num, 0) if val_num.is_integer() else fmt_num(val_num, 2)
        )
    except (ValueError, TypeError):
        val_str_formatted = str(val) if val is not None else ""

    val_str = f"{val_str_formatted} {unit_fixed}".strip()
    rows_to_render.append({"label": lbl, "value_str": val_str})

max_row = len(rows_to_render) + 1

# ==========================================
# 4. SATIR 1: BAŞLIK (SARI ALAN - A1:B1)
# ==========================================
ws.row_dimensions[1].height = 20
ws.merge_cells("A1:B1")

cell_a1 = ws["A1"]
cell_a1.value = header_title
cell_a1.font = font_header
cell_a1.alignment = Alignment(horizontal="center", vertical="center")

cell_b1 = ws["B1"]

cell_a1.fill = FILL_HEADER_YELLOW
cell_b1.fill = FILL_HEADER_YELLOW

cell_a1.border = Border(
    top=double_black_side, left=double_black_side, bottom=double_black_side
)
cell_b1.border = Border(
    top=double_black_side, right=double_black_side, bottom=double_black_side
)

# ==========================================
# 5. VERİ SATIRLARI DÖNGÜSÜ (Sadece A ve B)
# ==========================================
current_row = 2

for row_data in rows_to_render:
    ws.row_dimensions[current_row].height = 19
    is_last_row = current_row == max_row

    # A Kolonu: Etiket (Gri Dolgu)
    cell_a = ws.cell(row=current_row, column=1, value=row_data["label"])
    cell_a.fill = FILL_LABEL_GREY
    cell_a.font = font_label
    cell_a.alignment = Alignment(horizontal="left", vertical="center")

    # B Kolonu: Değer (Beyaz Dolgu)
    cell_b = ws.cell(row=current_row, column=2, value=row_data["value_str"])
    cell_b.fill = FILL_WHITE
    cell_b.font = font_value
    cell_b.alignment = Alignment(horizontal="left", vertical="center")

    b_bottom_a = double_black_side if is_last_row else None
    b_bottom_b = double_black_side if is_last_row else thin_grey_side

    cell_a.border = Border(left=double_black_side, bottom=b_bottom_a)
    cell_b.border = Border(right=double_black_side, bottom=b_bottom_b)

    current_row += 1

# ==========================================
# 6. SÜTUN GENİŞLİKLERİ VE KAYDETME (xlsx1'e)
# ==========================================
ws.column_dimensions["A"].width = 58
ws.column_dimensions["B"].width = 32

wb.save(OUTPUT_FILE)
print(f"🚀 Excel Kapak Tablosu Sıfırdan Çizilerek Kaydedildi: {OUTPUT_FILE}")
