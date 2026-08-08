import json
import os
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ==========================================
# 1. DOSYA & JSON VERİ OKUMA
# ==========================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

OUTPUT_DIR = os.path.join(BASE_DIR, "..", "tables", "xlsx1")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "on_yillik_maliyet.xlsx")

os.makedirs(OUTPUT_DIR, exist_ok=True)

JSON_PATH = os.path.join(BASE_DIR, "..", "..", "formData.json")

if not os.path.exists(JSON_PATH):
    raise FileNotFoundError(f"Hata: {JSON_PATH} dosyası bulunamadı!")

with open(JSON_PATH, "r", encoding="utf-8") as f:
    form_data = json.load(f)

# --- JSON YAPISINDAN FORMATLI STRING VERİLERİ ÇEKME ---
customer_info = form_data.get("customerInfo", {})
data_obj = form_data.get("tables", {}).get("onyillikmaliyettablosu", {})

selected_system = data_obj.get(
    "selectedSystem", form_data.get("selectedSystem", "mbbr")
)
is_mbbr = selected_system.lower() == "mbbr"
is_foreign = customer_info.get("teklifDili") == "Yabancı"

# React tarafında oluşturduğumuz hazır string verileri alıyoruz
planet_rendered = data_obj.get("planetRendered", {})
target_rendered = data_obj.get("altSystemRendered", {})
rendered_summary = data_obj.get("renderedSummary", {})

# --- SATIR VERİLERİNİ HAZIRLAMA (rows_data) ---
rows_data = []

# PlanetDISK String Verileri
if planet_rendered:
    rows_data.append(
        {
            "system": "PlanetDISK®",
            "capex": planet_rendered.get("capex", "-"),
            "energy": planet_rendered.get("energy", "-"),
            "operator": planet_rendered.get("operator", "-"),
            "maint": planet_rendered.get("maintenance", "-"),
        }
    )

# Karşılaştırılan Sistem String Verileri
if target_rendered:
    system_label = "MBBR" if is_mbbr else "Aktif Çamur"
    rows_data.append(
        {
            "system": system_label,
            "capex": target_rendered.get("capex", "-"),
            "energy": target_rendered.get("energy", "-"),
            "operator": target_rendered.get("operator", "-"),
            "maint": target_rendered.get("maintenance", "-"),
        }
    )

# --- ALT VURGU METNİ (Footer Gain) ---
footer_gain_text = rendered_summary.get("excelSavingsText")
if not footer_gain_text:
    savings_str = data_obj.get("totalSavings10YConverted", "0 €")
    if is_foreign:
        footer_gain_text = f"10 Years Total Savings with PlanetDISK®: {savings_str}"
    else:
        footer_gain_text = f"PlanetDISK® ile 10 Yıllık Toplam Kazanç: {savings_str}"

# ==========================================
# 2. EXCEL SIFIRDAN HAZIRLIK
# ==========================================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "TcoAnaliz"
ws.views.sheetView[0].showGridLines = True

# --- STİLLER VE FONT TANIMLARI ---
FILL_HEADER_YELLOW = PatternFill(
    start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"
)
FILL_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

font_title = Font(name="Calibri", size=9, bold=True)
font_bold = Font(name="Calibri", size=9, bold=True)
font_regular = Font(name="Calibri", size=9, bold=False)
font_red_footer = Font(name="Calibri", size=9, bold=True, italic=True, color="FF0000")

thin_s = Side(border_style="thin", color="000000")
double_s = Side(border_style="double", color="000000")

# ==========================================
# 3. SATIR 1: ÜST SARI BAŞLIK (A1:E1)
# ==========================================
ws.row_dimensions[1].height = 22
ws.merge_cells("A1:E1")

if is_foreign:
    sys_title = "MBBR SYSTEM" if is_mbbr else "ACTIVATED SLUDGE SYSTEM"
    title_text = f"PlanetDISK® DBD TECHNOLOGY vs {sys_title} 10 YEARS TOTAL COST OF OWNERSHIP (TCO) ANALYSIS"
else:
    sys_title = "MBBR SİSTEMİ" if is_mbbr else "AKTİF ÇAMUR SİSTEMİ"
    title_text = f"PlanetDISK® DBD TEKNOLOJİSİ İLE {sys_title} 10 YILLIK EKONOMİK ÖMÜR VEYA YATIRIM GERİ DÖNÜŞÜM (TCO) ANALİZİ"

cell_title = ws["A1"]
cell_title.value = title_text
cell_title.font = font_title
cell_title.alignment = Alignment(horizontal="center", vertical="center")

for c in range(1, 6):
    ws.cell(row=1, column=c).fill = FILL_HEADER_YELLOW
    ws.cell(row=1, column=c).border = Border(
        top=double_s,
        bottom=double_s,
        left=double_s if c == 1 else None,
        right=double_s if c == 5 else None,
    )

# ==========================================
# 4. SATIR 2: KOLON BAŞLIKLARI (A2:E2)
# ==========================================
ws.row_dimensions[2].height = 28

headers = [
    "",
    "Initial Investment Cost (Excl. Civil Works)"
    if is_foreign
    else "İlk Yatırım Maliyeti (İnşaat Hariç)",
    "Yearly Energy Cost" if is_foreign else "Yıllık Enerji Maliyeti",
    "Yearly Operator Cost" if is_foreign else "Yıllık Operatör Maliyeti",
    "Yearly Maintenance & Spare Parts Cost"
    if is_foreign
    else "Yıllık Bakım ve Yedek Parça Maliyeti",
]

for col_idx in range(1, 6):
    cell = ws.cell(row=2, column=col_idx, value=headers[col_idx - 1])
    cell.font = font_bold
    cell.fill = FILL_HEADER_YELLOW
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        top=thin_s,
        bottom=double_s,
        left=double_s if col_idx == 1 else thin_s,
        right=double_s if col_idx == 5 else thin_s,
    )

# ==========================================
# 5. SATIR 3-4: TABLO VERİ SATIRLARI (STRING OLARAK EKLENİR)
# ==========================================
current_row = 3

for row_item in rows_data:
    ws.row_dimensions[current_row].height = 24

    system_label = row_item.get("system", "")
    capex_val = row_item.get("capex", "")
    energy_val = row_item.get("energy", "")
    op_val = row_item.get("operator", "")
    maint_val = row_item.get("maint", "")

    c1 = ws.cell(row=current_row, column=1, value=system_label)
    c2 = ws.cell(row=current_row, column=2, value=capex_val)
    c3 = ws.cell(row=current_row, column=3, value=energy_val)
    c4 = ws.cell(row=current_row, column=4, value=op_val)
    c5 = ws.cell(row=current_row, column=5, value=maint_val)

    c1.alignment = Alignment(horizontal="center", vertical="center")
    c1.font = font_bold

    # Hücrelere direkt string eklendiği için sayı biçimlendirme (number_format) kaldırıldı
    for cell_obj in [c2, c3, c4, c5]:
        cell_obj.alignment = Alignment(horizontal="center", vertical="center")
        cell_obj.font = font_regular

    for c in range(1, 6):
        c_cell = ws.cell(row=current_row, column=c)
        c_cell.fill = FILL_WHITE
        c_cell.border = Border(
            top=thin_s,
            bottom=thin_s if current_row == 3 else double_s,
            left=double_s if c == 1 else thin_s,
            right=double_s if c == 5 else thin_s,
        )

    current_row += 1

# ==========================================
# 6. SATIR 5: EN ALT KIRMIZI VURGU SATIRI
# ==========================================
ws.row_dimensions[current_row].height = 24
ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)

for col in range(1, 6):
    cell = ws.cell(row=current_row, column=col)
    cell.font = font_red_footer
    cell.fill = FILL_HEADER_YELLOW
    cell.border = Border(
        top=double_s,
        bottom=double_s,
        left=double_s if col == 1 else None,
        right=double_s if col == 5 else None,
    )

footer_cell = ws.cell(row=current_row, column=1, value=footer_gain_text)
footer_cell.alignment = Alignment(horizontal="center", vertical="center")

# ==========================================
# 7. SÜTUN GENİŞLİKLERİ VE KAYDETME
# ==========================================
ws.column_dimensions["A"].width = 28
ws.column_dimensions["B"].width = 22
ws.column_dimensions["C"].width = 20
ws.column_dimensions["D"].width = 20
ws.column_dimensions["E"].width = 22

wb.save(OUTPUT_FILE)
print(f"🚀 Excel TCO Analiz Tablosu Güncellendi: {OUTPUT_FILE}")
