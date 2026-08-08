import json
import os
import openpyxl

# ==========================================
# 1. DOSYA & JSON VERİ OKUMA (GÜNCELLENDİ)
# ==========================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# 📂 Okunacak referans/şablon dosya (xlsx0)
SOURCE_FILE = os.path.join(BASE_DIR, "..", "tables", "xlsx0", "amortisman.xlsx")

# 💾 Kaydedilecek hedef dosya ve klasör (xlsx1)
OUTPUT_DIR = os.path.join(BASE_DIR, "..", "tables", "xlsx1")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "amortisman.xlsx")

# Target klasör (xlsx1) yoksa otomatik oluştur
os.makedirs(OUTPUT_DIR, exist_ok=True)

JSON_PATH = os.path.join(BASE_DIR, "..", "..", "formData.json")

if not os.path.exists(JSON_PATH):
    raise FileNotFoundError(f"Hata: {JSON_PATH} dosyası bulunamadı!")

with open(JSON_PATH, "r", encoding="utf-8") as f:
    form_data = json.load(f)

# Tablo Verisi
amort_data = form_data.get("tables", {}).get("amortisman", {})

# Customer Info Yapısı
customer_info = form_data.get("customerInfo", {})

# ==========================================
# 🎯 DİL, PARA BİRİMİ VE BİRİM SİSTEMİ TESPİTİ
# ==========================================
# Sadece teklifDili kontrol edilir
raw_teklif_dili = (
    customer_info.get("teklifDili") or amort_data.get("teklifDili") or "Yerli"
)
is_turkish = str(raw_teklif_dili).strip().lower() == "yerli"

currency_code = str(
    customer_info.get("currency") or amort_data.get("currency") or "EUR"
).upper()
unit_system = str(
    customer_info.get("unitSystem") or amort_data.get("unitSystem") or "Metric"
).upper()
is_us = unit_system == "US"

exchange_rate = float(
    customer_info.get("exchangeRate") or amort_data.get("exchangeRate") or 1.0
)

if currency_code == "USD":
    CURRENCY_SYMBOL = "$"
elif currency_code == "TRY":
    CURRENCY_SYMBOL = "₺"
else:
    CURRENCY_SYMBOL = "€"


def fmt_num(val, decimals=2):
    """Sayıyı teklifDili'ne göre biçimlendirir."""
    if val is None or val == "":
        return "0"
    try:
        val_float = float(val)
        if decimals == 0:
            val_float = round(val_float)
    except (ValueError, TypeError):
        return str(val)

    formatted = f"{val_float:,.{decimals}f}" if decimals > 0 else f"{val_float:,}"

    if is_turkish:
        # Türkçe modunda binlik ayracı nokta, ondalık ayracı virgül
        formatted = formatted.replace(",", "X").replace(".", ",").replace("X", ".")

    return formatted


# ==========================================
# 2. HESAPLAMALAR VE RENDER DEĞERLERİ
# ==========================================
rendered_metrics = amort_data.get("renderedMetrics", {})
rendered_summary = amort_data.get("renderedSummary", {})

# Ham girdiler
daily_usage_raw = float(amort_data.get("dailyUsage", 70) or 70)
active_months = int(amort_data.get("activeMonths", 7) or 7)
water_price_raw = float(amort_data.get("waterPrice", 1.59) or 1.59)
plant_cost_raw = float(amort_data.get("plantCost", 327457) or 327457)
annual_opex_raw = float(amort_data.get("annualOpex", 0) or 0)

# Birim / Kur Dönüşümlü Hesaplamalar
calc_daily_usage = daily_usage_raw * 264.172 if is_us else daily_usage_raw
calc_monthly_usage = calc_daily_usage * 30
calc_yearly_usage = calc_monthly_usage * active_months

calc_water_price = (
    (water_price_raw / 264.172) * exchange_rate
    if is_us
    else water_price_raw * exchange_rate
)
calc_yearly_water_cost = calc_yearly_usage * calc_water_price

calc_plant_cost = plant_cost_raw * exchange_rate
calc_annual_opex = annual_opex_raw * exchange_rate

calc_net_saving = calc_yearly_water_cost - calc_annual_opex
calc_roi_years = calc_plant_cost / calc_net_saving if calc_net_saving > 0 else 0
calc_roi_months = calc_roi_years * 12
calc_exact_year_round = int(round(calc_roi_years))

# Final Render Değerleri
daily_usage = float(rendered_metrics.get("dailyUsage", calc_daily_usage))
monthly_usage = float(rendered_metrics.get("monthlyUsage", calc_monthly_usage))
yearly_usage = float(rendered_metrics.get("yearlyUsage", calc_yearly_usage))
water_price = float(rendered_metrics.get("waterPrice", calc_water_price))
yearly_water_cost = float(
    rendered_metrics.get("yearlyWaterCost", calc_yearly_water_cost)
)
plant_cost = float(rendered_metrics.get("plantCost", calc_plant_cost))
annual_opex = float(rendered_metrics.get("annualOpex", calc_annual_opex))

roi_years = float(rendered_metrics.get("roiYears", calc_roi_years))
roi_months = float(rendered_metrics.get("roiMonths", calc_roi_months))
exact_year_round = int(rendered_metrics.get("exactYearRound", calc_exact_year_round))

# ==========================================
# 3. EXCEL DOSYASINI YÜKLEME (xlsx0'dan)
# ==========================================
if not os.path.exists(SOURCE_FILE):
    raise FileNotFoundError(f"Hata: Şablon dosya bulunamadı! Yol: {SOURCE_FILE}")

# Referans excel dosyasını okuyoruz
wb = openpyxl.load_workbook(SOURCE_FILE)

if "Amortisman" in wb.sheetnames:
    ws = wb["Amortisman"]
elif "AMORTIZATION" in wb.sheetnames:
    ws = wb["AMORTIZATION"]
else:
    ws = wb.active

ws.views.sheetView[0].showGridLines = True

# ==========================================
# 4. BAŞLIK VE DİLE GÖRE TABLO DOLDURMA
# ==========================================
if is_turkish:
    # 🇹🇷 TÜRKÇE MODU (teklifDili == "Yerli")
    ws["A1"] = "YATIRIMIN GERİ DÖNÜŞ SÜRESİ (AMORTİSMAN) TABLOSU"

    ws["A2"] = "SULAMA AMAÇLI ŞEBEKE SUYUNUN KULLANILMASI DURUMUNDA"  # A2:B2 birleşik
    ws["C2"] = "Günlük su kullanımı"
    ws["D2"] = "Aylık su kullanımı"
    ws["E2"] = f"Yılda {active_months} ay su kullanımı"
    ws["F2"] = "Şebeke suyu birim fiyatı"
    ws["G2"] = "Toplam yıllık su bedeli"

    ws["A3"] = "BİRİM"  # A3:B3 birleşik
    ws["C3"] = "GPD" if is_us else "m³/gün"
    ws["D3"] = "Gallons/month" if is_us else "m³/ay"
    ws["E3"] = "Gallons/year" if is_us else "m³/yıl"
    ws["F3"] = f"{CURRENCY_SYMBOL}/gal" if is_us else f"{CURRENCY_SYMBOL}/m³"
    ws["G3"] = f"{CURRENCY_SYMBOL} / yıl"

    ws["A4"] = "DEĞER"  # A4:B4 birleşik

    ws["A5"] = "SULAMA AMAÇLI EVSEL ATIKSU ARITMA TESİSİNDEN ÇIKAN SU KULLANILIRSA"
    ws["B5"] = "Atıksu Arıtma Tesisinin Yaklaşık Maliyeti"  # B5:C5 birleşik
    ws["D5"] = "ATIKSU ARITMA TESİSİNİN AMORTİ ETME SÜRESİ"  # D5:G5 birleşik

    ws["A6"] = "BİRİM"
    ws["B6"] = CURRENCY_SYMBOL  # B6:C6 birleşik
    ws["D6"] = "Yıl"  # D6:G6 birleşik

    ws["A7"] = "DEĞER"

    ws["A8"] = "Sistem kendisini ancak tam"
    ws["B8"] = int(round(roi_months))
    ws["C8"] = "AYDA geri döndürebilmektedir."  # C8:G8 birleşik
    ws["A10"] = (
        f"⚠️ Not: Bu süreye her yıl güncellenen amortisman tablosundaki işletme giderleri ({fmt_num(annual_opex, 0)} {CURRENCY_SYMBOL}/yıl) dahil edilerek hesaplama yapılmıştır."
    )

else:
    # 🇬🇧 İNGİLİZCE MODU (teklifDili == "Yabancı")
    ws["A1"] = "AMORTIZATION TABLE"

    ws["A2"] = "IF THE IRRIGATION WATER IS SUPPLIED FROM MUNICIPAL WATER"  # A2:B2
    ws["C2"] = "Daily water required"
    ws["D2"] = "Monthly water required"
    ws["E2"] = "Yearly water required"
    ws["F2"] = "Unit Price of Municipal Water"
    ws["G2"] = "Yearly Total Water Use Cost"

    ws["A3"] = "UNIT"  # A3:B3
    ws["C3"] = "GPD" if is_us else "m³/day"
    ws["D3"] = "Gallons/month" if is_us else "m³/month"
    ws["E3"] = "Gallons/year" if is_us else "m³/year"
    ws["F3"] = f"{CURRENCY_SYMBOL}/gal" if is_us else f"{CURRENCY_SYMBOL}"
    ws["G3"] = f"{CURRENCY_SYMBOL}/year"

    ws["A4"] = "VALUE"  # A4:B4

    ws["A5"] = "IF THE IRRIGATION WATER IS SUPPLIED FROM WWTP"
    ws["B5"] = "WWTP CAPEX"  # B5:C5
    ws["D5"] = "Amortization Time"  # D5:G5

    ws["A6"] = "UNIT"
    ws["B6"] = CURRENCY_SYMBOL  # B6:C6
    ws["D6"] = "Year"  # D6:G6

    ws["A7"] = "VALUE"

    ws["A8"] = "In"
    ws["B8"] = int(round(roi_months))
    ws["C8"] = "Months, the WWTP is amortizing itself."  # C8:G8
    ws["A10"] = (
        f'*** "Unit Price of Municipal Water = {fmt_num(water_price, 4)} {CURRENCY_SYMBOL}/{"gal" if is_us else "m³"}" is given for comparison purposes.'
    )

# ==========================================
# 5. SAYISAL DEĞERLERİ YAZMA
# ==========================================
ws["C4"] = fmt_num(daily_usage, 2)
ws["D4"] = fmt_num(monthly_usage, 0)
ws["E4"] = fmt_num(yearly_usage, 0)
ws["F4"] = fmt_num(water_price, 4)
ws["G4"] = fmt_num(yearly_water_cost, 0)

ws["B7"] = fmt_num(plant_cost, 0)
ws["D7"] = fmt_num(roi_years, 2)

# ==========================================
# 6. KAYDETME (xlsx1'e)
# ==========================================
# openpyxl yüklü çalışma kitabını orijinal dosyanın üstüne değil,
# yeni oluşturduğumuz OUTPUT_FILE (xlsx1) yoluna kaydeder.
wb.save(OUTPUT_FILE)

print(f"✅ Referans Alınan Şablon : {SOURCE_FILE}")
print(
    f"🚀 Oluşturulan Yeni Tablo : {OUTPUT_FILE} (Dil: {'Türkçe' if is_turkish else 'İngilizce'})"
)
