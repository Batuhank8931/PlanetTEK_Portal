import json
import os
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ==========================================
# 1. DOSYA & JSON VERİ OKUMA (GÜNCELLENDİ)
# ==========================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# 💾 Kaydedilecek hedef dosya ve klasör (xlsx1)
OUTPUT_DIR = os.path.join(BASE_DIR, "..", "tables", "xlsx1")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "opex.xlsx")

# Target klasör (xlsx1) yoksa otomatik oluştur
os.makedirs(OUTPUT_DIR, exist_ok=True)

JSON_PATH = os.path.join(BASE_DIR, "..", "..", "formData.json")

if not os.path.exists(JSON_PATH):
    raise FileNotFoundError(f"Hata: {JSON_PATH} dosyası bulunamadı!")

with open(JSON_PATH, "r", encoding="utf-8") as f:
    form_data = json.load(f)

customer_info = form_data.get("customerInfo", {})
opex_data_obj = form_data.get("tables", {}).get("opextablosu", {})

rows_data = opex_data_obj.get("rows", [])
total_opex_json = opex_data_obj.get("totalOpex", 0)

# --- DİL SEÇİMİ VE PARA BİRİMİ DİNAMİKLERİ ---
is_foreign = customer_info.get("teklifDili") == "Yabancı"
currency_code = customer_info.get("currency", "EUR").upper()

if currency_code == "USD":
    CURRENCY_SYMBOL = "$"
elif currency_code == "TRY":
    CURRENCY_SYMBOL = "₺"
else:
    CURRENCY_SYMBOL = "€"


def fmt_num(val, decimals=0):
    """Sayıyı teklifDili'ne göre formatlar."""
    try:
        val = float(val)
    except (ValueError, TypeError):
        val = 0.0

    formatted = f"{val:,.{decimals}f}"

    if not is_foreign:
        formatted = formatted.replace(",", "X").replace(".", ",").replace("X", ".")

    return formatted


LANG = {
    "title": (
        "OPERATION EXPENDITURE - OPEX"
        if is_foreign
        else "YILLIK İŞLETME GİDERLERİ - OPEX"
    ),
    "col_desc": "Description" if is_foreign else "Giderlerin Tanımları",
    "col_price": "Total Price" if is_foreign else "Toplam Fiyat",
    "grand_total": "GRAND TOTAL" if is_foreign else "GENEL TOPLAM",
    "year_unit": (
        f"{CURRENCY_SYMBOL} /year" if is_foreign else f"{CURRENCY_SYMBOL} /yıl"
    ),
    "rows": {
        "enerji_gideri": (
            "Energy Operation Cost" if is_foreign else "Enerji Giderleri"
        ),
        "sarf_gideri": (
            "Consumables and Maintenance Cost"
            if is_foreign
            else "Sarf Malzemesi ve Bakım Giderleri"
        ),
    },
}

# ==========================================
# 2. EXCEL SIFIRDAN HAZIRLIK
# ==========================================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Opex"
ws.views.sheetView[0].showGridLines = True

# --- STİLLER ---
FILL_YELLOW = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
FILL_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

font_main_title = Font(name="Calibri", size=9, bold=True, italic=True)
font_bold = Font(name="Calibri", size=9, bold=True)
font_regular = Font(name="Calibri", size=9, bold=False)

thin_s = Side(border_style="thin", color="000000")
double_s = Side(border_style="double", color="000000")

# ==========================================
# 3. ÜST SARI BAŞLIK ALANI (A1:B2)
# ==========================================
# Satır 1: Ana Başlık (A1:B1)
ws.row_dimensions[1].height = 20
ws.merge_cells("A1:B1")
ws["A1"] = LANG["title"]
ws["A1"].font = font_main_title
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

ws["A1"].fill = FILL_YELLOW
ws["B1"].fill = FILL_YELLOW
ws["A1"].border = Border(top=double_s, bottom=double_s, left=double_s)
ws["B1"].border = Border(top=double_s, bottom=double_s, right=double_s)

# Satır 2: Kolon Başlıkları (A2:B2)
ws.row_dimensions[2].height = 18
ws["A2"] = LANG["col_desc"]
ws["A2"].font = font_bold
ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
ws["A2"].fill = FILL_YELLOW
ws["A2"].border = Border(top=thin_s, bottom=double_s, left=double_s, right=thin_s)

ws["B2"] = LANG["col_price"]
ws["B2"].font = font_bold
ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
ws["B2"].fill = FILL_YELLOW
ws["B2"].border = Border(top=thin_s, bottom=double_s, left=thin_s, right=double_s)

# ==========================================
# 4. SABİT VERİ SATIRLARI (Satır 3 ve Satır 4)
# ==========================================
val_enerji = 0.0
val_sarf = 0.0

for r_item in rows_data:
    r_id = r_item.get("id", "")
    r_val = float(r_item.get("value", 0) or 0)
    if r_id == "enerji_gideri":
        val_enerji = r_val
    elif r_id == "sarf_gideri":
        val_sarf = r_val

# Satır 3: Enerji Giderleri
ws.row_dimensions[3].height = 18
ws["A3"] = f"  {LANG['rows']['enerji_gideri']}"
ws["A3"].font = font_bold
ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
ws["A3"].fill = FILL_WHITE
ws["A3"].border = Border(left=double_s, right=thin_s, top=None, bottom=None)

ws["B3"] = f"{fmt_num(val_enerji, 0)} {LANG['year_unit']}"
ws["B3"].font = font_regular
ws["B3"].alignment = Alignment(horizontal="center", vertical="center")
ws["B3"].fill = FILL_WHITE
ws["B3"].border = Border(left=thin_s, right=double_s, top=None, bottom=None)

# Satır 4: Sarf Malzemesi ve Bakım Giderleri
ws.row_dimensions[4].height = 18
ws["A4"] = f"  {LANG['rows']['sarf_gideri']}"
ws["A4"].font = font_bold
ws["A4"].alignment = Alignment(horizontal="left", vertical="center")
ws["A4"].fill = FILL_WHITE
ws["A4"].border = Border(left=double_s, right=thin_s, top=None, bottom=None)

ws["B4"] = f"{fmt_num(val_sarf, 0)} {LANG['year_unit']}"
ws["B4"].font = font_regular
ws["B4"].alignment = Alignment(horizontal="center", vertical="center")
ws["B4"].fill = FILL_WHITE
ws["B4"].border = Border(left=thin_s, right=double_s, top=None, bottom=None)

# ==========================================
# 5. GENEL TOPLAM SATIRI (Satır 5)
# ==========================================
ws.row_dimensions[5].height = 20
ws["A5"] = LANG["grand_total"]
ws["A5"].font = font_bold
ws["A5"].alignment = Alignment(horizontal="right", vertical="center")
ws["A5"].fill = FILL_YELLOW
ws["A5"].border = Border(top=thin_s, bottom=double_s, left=double_s, right=thin_s)

final_total = total_opex_json if total_opex_json > 0 else (val_enerji + val_sarf)
ws["B5"] = f"{fmt_num(final_total, 0)} {LANG['year_unit']}"
ws["B5"].font = font_bold
ws["B5"].alignment = Alignment(horizontal="center", vertical="center")
ws["B5"].fill = FILL_YELLOW
ws["B5"].border = Border(top=thin_s, bottom=double_s, left=thin_s, right=double_s)

# ==========================================
# 6. KOLON GENİŞLİKLERİ VE KAYDETME (xlsx1'e)
# ==========================================
ws.column_dimensions["A"].width = 42
ws.column_dimensions["B"].width = 25

wb.save(OUTPUT_FILE)
print(f"🚀 Excel OPEX Tablosu Sıfırdan Çizilerek Kaydedildi: {OUTPUT_FILE}")
