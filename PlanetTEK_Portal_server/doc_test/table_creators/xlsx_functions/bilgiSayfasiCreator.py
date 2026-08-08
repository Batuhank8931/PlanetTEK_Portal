import json
import os
import openpyxl
from openpyxl.styles import Alignment, Font

# ==========================================
# 1. DOSYA & JSON VERİ OKUMA
# ==========================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# 📂 Okunacak referans/şablon dosya (xlsx0)
SOURCE_FILE = os.path.join(BASE_DIR, "..", "tables", "xlsx0", "bilgi_sayfasi.xlsx")

# 💾 Kaydedilecek hedef dosya ve klasör (xlsx1)
OUTPUT_DIR = os.path.join(BASE_DIR, "..", "tables", "xlsx1")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "bilgi_sayfasi.xlsx")

os.makedirs(OUTPUT_DIR, exist_ok=True)

JSON_PATH = os.path.join(BASE_DIR, "..", "..", "formData.json")

if not os.path.exists(JSON_PATH):
    raise FileNotFoundError(f"Hata: {JSON_PATH} dosyası bulunamadı!")

with open(JSON_PATH, "r", encoding="utf-8") as f:
    form_data = json.load(f)

customer_info = form_data.get("customerInfo", {})
bilgi_data = form_data.get("tables", {}).get("bilgisayfasitablosu", {})

if not bilgi_data:
    bilgi_data = form_data.get("tables", {}).get("bilgisayfasi", {})

# ==========================================
# 🎯 DİL VE DİNAMİK METİN KONTROLÜ
# ==========================================
raw_teklif_dili = (
    customer_info.get("teklifDili") or bilgi_data.get("teklifDili") or "Yerli"
)
is_turkish = str(raw_teklif_dili).strip().lower() == "yerli"

title1 = bilgi_data.get("title1", customer_info.get("ticari_unvan", ""))
title2 = bilgi_data.get("title2", "")
title3 = bilgi_data.get("title3", "")

details_header = bilgi_data.get(
    "detailsHeader", "PROJE DETAYLARI" if is_turkish else "PROJECT INFO"
)
project_details = bilgi_data.get("projectDetails", [])
note_text = bilgi_data.get("noteText", "")

source_header = bilgi_data.get(
    "sourceHeader", "Atıksu Kaynağı" if is_turkish else "Wastewater Source"
)
source_text = bilgi_data.get("sourceText", "")

system_header = bilgi_data.get(
    "systemHeader", "Önerilen Sistem" if is_turkish else "Proposed System"
)
system_text_raw = bilgi_data.get("systemText", "")

calc_header = bilgi_data.get(
    "calcHeader",
    ("Disk Yüzey Alanı Hesaplaması" if is_turkish else "Disk Surface Area Calculation"),
)
calc_text = bilgi_data.get("calcText", "")

# ==========================================
# 2. TEMİZ ŞABLONU YÜKLEME (xlsx0'dan)
# ==========================================
if not os.path.exists(SOURCE_FILE):
    raise FileNotFoundError(f"Hata: Şablon dosyası bulunamadı! ({SOURCE_FILE})")

wb = openpyxl.load_workbook(SOURCE_FILE)

if "Bilgi Sayfası" in wb.sheetnames:
    ws = wb["Bilgi Sayfası"]
elif "BilgiSayfasi" in wb.sheetnames:
    ws = wb["BilgiSayfasi"]
else:
    ws = wb.active

ws.views.sheetView[0].showGridLines = True

# ==========================================
# 3. HÜCRELERE YAZMA VE BİÇİMLENDİRME
# ==========================================

# 1. Üst Başlıklar
ws["A1"] = title1
ws["A2"] = title2
ws["A3"] = title3

# 2. Proje Detayları Başlığı
ws["A5"] = details_header

# 3. Proje Detay Satırları (B6:C15 Sabit Şablon Alanı)
for r in range(6, 16):
    ws[f"B{r}"] = None
    ws[f"C{r}"] = None

current_row = 6
for item in project_details:
    if current_row > 15:
        break
    label = item.get("label", "")
    value = item.get("value", "")

    ws[f"B{current_row}"] = label
    ws[f"C{current_row}"] = value
    current_row += 1

# 4. Dipnot Metni (A17)
ws["A17"] = note_text

# 5. Atıksu Kaynağı Bölümü (A19 & A20)
ws["A19"] = source_header
ws["A20"] = source_text
# A20 için şablon biçimlendirmesini koru (10pt Bold, Left-Top, Indent=2, WrapText=True)
ws["A20"].font = Font(name="Calibri", size=10, bold=True)
ws["A20"].alignment = Alignment(
    horizontal="left", vertical="top", indent=2, wrap_text=True
)

# 6. Önerilen Sistem Bölümü (A22 & A23:A28)
ws["A22"] = system_header
ws["A22"].font = Font(name="Calibri", size=12, bold=True)
ws["A22"].alignment = Alignment(horizontal="center", vertical="center")

# Dinamik liste ayrıştırma
raw_lines = [line.strip() for line in str(system_text_raw).split("\n") if line.strip()]

system_lines = []
for line in raw_lines:
    clean_line = line.lstrip("·").lstrip("-").strip()
    system_lines.append(f"· {clean_line}")

# A23-A28 aralığını temizle
for idx in range(23, 29):
    ws[f"A{idx}"] = None

sys_row = 23
for line in system_lines:
    if sys_row <= 28:
        cell = ws[f"A{sys_row}"]
        cell.value = line
        # Şablondaki font (10pt regular) ve hizalamayı (left, center, indent=2) koru
        cell.font = Font(name="Calibri", size=10, bold=False)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        sys_row += 1

# 7. Disk Yüzey Alanı Hesaplaması Bölümü (A31 Başlık & A32 Değer)
# ⚠️ DÜZELTME: Şablonda başlık A31 hücresindedir (12pt Bold Center)
ws["A31"] = calc_header
ws["A31"].font = Font(name="Calibri", size=12, bold=True)
ws["A31"].alignment = Alignment(horizontal="center", vertical="center")

# Hesaplama metni A32 hücresine yazılır (10pt Regular Center)
ws["A32"] = calc_text
ws["A32"].font = Font(name="Calibri", size=10, bold=False)
ws["A32"].alignment = Alignment(horizontal="center", vertical="center")

# ==========================================
# 4. KAYDETME (xlsx1'e)
# ==========================================
wb.save(OUTPUT_FILE)

print(f"✅ Referans Alınan Şablon : {SOURCE_FILE}")
print(f"🚀 Oluşturulan Yeni Tablo : {OUTPUT_FILE}")
