import json
import os
import openpyxl

# ==========================================
# 1. DOSYA & JSON VERİ OKUMA (GÜNCELLENDİ)
# ==========================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# 📂 Okunacak referans/şablon dosya (xlsx0)
SOURCE_FILE = os.path.join(BASE_DIR, "..", "tables", "xlsx0", "karbon_ayakizi.xlsx")

# 💾 Kaydedilecek hedef dosya ve klasör (xlsx1)
OUTPUT_DIR = os.path.join(BASE_DIR, "..", "tables", "xlsx1")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "karbon_ayakizi.xlsx")

# Target klasör (xlsx1) yoksa otomatik oluştur
os.makedirs(OUTPUT_DIR, exist_ok=True)

JSON_PATH = os.path.join(BASE_DIR, "..", "..", "formData.json")

if not os.path.exists(JSON_PATH):
    raise FileNotFoundError(f"Hata: {JSON_PATH} dosyası bulunamadı!")

with open(JSON_PATH, "r", encoding="utf-8") as f:
    form_data = json.load(f)

customer_info = form_data.get("customerInfo", {})
karbon_data = form_data.get("tables", {}).get("karbonayakizitablosu", {})

if not karbon_data:
    karbon_data = form_data.get("tables", {}).get("karbonayakizi", {})

# Dil ve Birim Sistemi
is_foreign = customer_info.get("teklifDili") == "Yabancı"
unit_system = customer_info.get("unitSystem", "Metric")
is_us = unit_system == "US"

# Sistem Seçimi (Aktif Çamur veya MBBR)
selected_system = karbon_data.get("selectedSystem", "aktif_camur")
is_mbbr = str(selected_system).lower() == "mbbr"

# ==========================================
# 2. HESAPLAMA VE DEĞER ELDE ETME
# ==========================================
# 1. CO2 Faktörü
co2_factor = float(karbon_data.get("co2Factor", 0.43) or 0.43)

# 2. Günlük Tüketimler (kWh/gün)
planet_daily_kwh = float(karbon_data.get("currentPlanetDailyKwh", 8.0) or 8.0)
alt_daily_kwh = float(karbon_data.get("currentAltDailyKwh", 51.1) or 51.1)

# 3. Yıllık Tüketimler (kWh/yıl)
planet_yearly_kwh = float(
    karbon_data.get("planetYearlyKwh", planet_daily_kwh * 365)
    or (planet_daily_kwh * 365)
)
alt_yearly_kwh = float(
    karbon_data.get("altYearlyKwh", alt_daily_kwh * 365) or (alt_daily_kwh * 365)
)

# 4. Karbon Ayak İzi (Ton/yıl)
ton_to_us_ton = 1.10231
calc_planet_co2 = (planet_yearly_kwh * co2_factor) / 1000.0
calc_alt_co2 = (alt_yearly_kwh * co2_factor) / 1000.0

if is_us:
    calc_planet_co2 *= ton_to_us_ton
    calc_alt_co2 *= ton_to_us_ton

planet_co2 = float(karbon_data.get("planetCo2", calc_planet_co2) or calc_planet_co2)
alt_co2 = float(karbon_data.get("altCo2", calc_alt_co2) or calc_alt_co2)

# 5. Ağaç Borcu / Eşdeğeri Hesabı
planet_tree_debt = (planet_co2 * 2000 / 48.5) if is_us else (planet_co2 * 1000 / 22)
alt_tree_debt = (alt_co2 * 2000 / 48.5) if is_us else (alt_co2 * 1000 / 22)


# ==========================================
# 3. SAYI FORMATLAMA YARDIMCISI
# ==========================================
def fmt_num(val, decimals=1):
    if val is None or val == "":
        return "0"
    try:
        val_float = float(val)
    except (ValueError, TypeError):
        return str(val)

    formatted = f"{val_float:,.{decimals}f}"

    if not is_foreign and not is_us:
        formatted = formatted.replace(",", "X").replace(".", ",").replace("X", ".")

    return formatted


# ==========================================
# 4. EXCEL DOSYASINI YÜKLEME VE DOLDURMA (xlsx0'dan)
# ==========================================
if not os.path.exists(SOURCE_FILE):
    raise FileNotFoundError(f"Hata: Şablon dosya bulunamadı! Yol: {SOURCE_FILE}")

wb = openpyxl.load_workbook(SOURCE_FILE)
ws = wb.active
ws.views.sheetView[0].showGridLines = True

# --- BAŞLIKLAR & ETİKETLER ---
alt_sys_code = "MBBR" if is_mbbr else "AS"
alt_sys_full = "MBBR System" if is_mbbr else "Activated Sludge System"

if is_foreign:
    ws["A1"] = "CARBON FOOTPRINT"
    ws["A3"] = "RBC System Energy Consumption"
    ws["A4"] = "RBC System Energy Consumption"
    ws["A5"] = "CO2 Emission Coefficient"
    ws["A6"] = "Annual Carbon Footprint"
    ws["A7"] = "Due to electricity consumption, there is an annual debt of"

    ws["D3"] = f"{alt_sys_code} System Energy Consumption"
    ws["D4"] = f"{alt_sys_code} System Energy Consumption"
    ws["D5"] = "CO2 Emission Coefficient"
    ws["D6"] = "Annual Carbon Footprint"
    ws["D7"] = "Due to electricity consumption, there is an annual debt of"
else:
    ws["A1"] = "KARBON AYAK İZİ"
    ws["A3"] = "DBD Sistemi Enerji Tüketimi"
    ws["A4"] = "DBD Sistemi Enerji Tüketimi"
    ws["A5"] = "CO2 Emisyon Faktörü"
    ws["A6"] = "Yıllık Karbon Ayak İzi"
    ws["A7"] = "Elektrik tüketimine bağlı yıllık ağaç borcu"

    ws["D3"] = f"{alt_sys_code} Sistemi Enerji Tüketimi"
    ws["D4"] = f"{alt_sys_code} Sistemi Enerji Tüketimi"
    ws["D5"] = "CO2 Emisyon Faktörü"
    ws["D6"] = "Yıllık Karbon Ayak İzi"
    ws["D7"] = "Elektrik tüketimine bağlı yıllık ağaç borcu"

# BİRİMLER (C VE F KOLONLARI)
kw_day_unit = "kw/day" if (is_foreign or is_us) else "kWh/gün"
kw_year_unit = "kW/year" if (is_foreign or is_us) else "kWh/yıl"
co2_unit = "kg/eMWh" if (is_us or is_foreign) else "kg CO₂/kWh"
ton_unit = "Ton/year" if is_us else ("ton/year" if is_foreign else "ton CO₂/yıl")
tree_unit = "trees to nature." if (is_foreign or is_us) else "Ağaç / yıl"

# ==========================================
# 5. SOL TARAF: RBC / PlanetDISK® SİSTEMİ (B & C Kolonları)
# ==========================================
ws["B3"] = fmt_num(planet_daily_kwh, 1)
ws["C3"] = kw_day_unit

ws["B4"] = fmt_num(round(planet_yearly_kwh), 0)
ws["C4"] = kw_year_unit

ws["B5"] = fmt_num(co2_factor, 2 if co2_factor < 1 else 3)
ws["C5"] = co2_unit

ws["B6"] = fmt_num(planet_co2, 1)
ws["C6"] = ton_unit

ws["B7"] = int(round(planet_tree_debt))
ws["C7"] = tree_unit

# ==========================================
# 6. SAĞ TARAF: KARSILASTIRILAN SİSTEM (E & F Kolonları)
# ==========================================
ws["E3"] = fmt_num(alt_daily_kwh, 1)
ws["F3"] = kw_day_unit

ws["E4"] = fmt_num(round(alt_yearly_kwh), 0)
ws["F4"] = kw_year_unit

ws["E5"] = fmt_num(co2_factor, 2 if co2_factor < 1 else 3)
ws["F5"] = co2_unit

ws["E6"] = fmt_num(alt_co2, 1)
ws["F6"] = ton_unit

ws["E7"] = int(round(alt_tree_debt))
ws["F7"] = tree_unit

# ==========================================
# 7. KAYDETME (xlsx1'e)
# ==========================================
wb.save(OUTPUT_FILE)

print(f"✅ Referans Alınan Şablon : {SOURCE_FILE}")
print(f"🚀 Oluşturulan Yeni Tablo : {OUTPUT_FILE}")
