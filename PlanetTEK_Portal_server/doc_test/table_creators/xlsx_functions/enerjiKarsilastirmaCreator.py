import json
import os
import openpyxl
from openpyxl.styles import PatternFill, Alignment

# ==========================================
# 1. DOSYA & JSON VERİ OKUMA
# ==========================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

SOURCE_FILE = os.path.join(
    BASE_DIR, "..", "tables", "xlsx0", "enerji_karsilastirma.xlsx"
)
OUTPUT_DIR = os.path.join(BASE_DIR, "..", "tables", "xlsx1")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "enerji_karsilastirma.xlsx")

os.makedirs(OUTPUT_DIR, exist_ok=True)

JSON_PATH = os.path.join(BASE_DIR, "..", "..", "formData.json")

if not os.path.exists(JSON_PATH):
    raise FileNotFoundError(f"Hata: {JSON_PATH} dosyası bulunamadı!")

with open(JSON_PATH, "r", encoding="utf-8") as f:
    form_data = json.load(f)

customer_info = form_data.get("customerInfo", {})
cmp_data = form_data.get("tables", {}).get("enerjikarsilastirmatablosu", {})

if not cmp_data:
    cmp_data = form_data.get("tables", {}).get("enerjikarsilastirma", {})

data_obj = cmp_data.get("data", {})
planet_raw = data_obj.get("planet", {})
blower_raw = data_obj.get("blower", {})
pump_raw = data_obj.get("pump", {})

planet_metrics = cmp_data.get("planetMetrics", {})
blower_metrics = cmp_data.get("blowerMetrics", {})
pump_metrics = cmp_data.get("pumpMetrics", {})

headers = cmp_data.get("headersAndLabels", {})
rendered_content = cmp_data.get("renderedTableContent", {})
summary_cards = rendered_content.get("summaryCards", {})

selected_system = cmp_data.get("selectedSystem", "aktif_camur")
is_mbbr = str(selected_system).lower() == "mbbr"

is_foreign = customer_info.get("teklifDili") == "Yabancı"
currency_code = customer_info.get("currency", "EUR").upper()
exchange_rate = float(customer_info.get("exchangeRate", 1.0) or 1.0)

if currency_code == "USD":
    CURRENCY_SYMBOL = "$"
elif currency_code == "TRY":
    CURRENCY_SYMBOL = "₺"
else:
    CURRENCY_SYMBOL = "€"


def fmt_num(val, decimals=2):
    if val is None or val == "":
        return "0"
    try:
        if isinstance(val, str):
            val = val.replace(",", "")
        val_float = float(val)
    except (ValueError, TypeError):
        return str(val)

    formatted = f"{val_float:,.{decimals}f}"
    if not is_foreign:
        formatted = formatted.replace(",", "X").replace(".", ",").replace("X", ".")
    return formatted


# ==========================================
# 2. ŞABLON YÜKLEME
# ==========================================
if not os.path.exists(SOURCE_FILE):
    raise FileNotFoundError(f"Hata: Şablon dosya bulunamadı! Yol: {SOURCE_FILE}")

wb = openpyxl.load_workbook(SOURCE_FILE)
ws = wb.active
ws.views.sheetView[0].showGridLines = True

FILL_GREEN_HEADER = PatternFill(
    start_color="92D05E", end_color="92D05E", fill_type="solid"
)
FILL_BLUE_HEADER = PatternFill(
    start_color="8EA9DB", end_color="8EA9DB", fill_type="solid"
)
sys_fill = FILL_BLUE_HEADER if is_mbbr else FILL_GREEN_HEADER


# ==========================================
# 3. ANA BAŞLIKLAR & SÜTUN BAŞLIKLARI
# ==========================================
alt_system_title = headers.get(
    "colHeaderAltSystem", ("MBBR SİSTEMİ" if is_mbbr else "AKTİF ÇAMUR SİSTEMİ")
)

ws["A1"] = headers.get(
    "panelTitle",
    f"PlanetDISK® DBD TEKNOLOJİSİ İLE {alt_system_title} ENERJİ KARŞILAŞTIRMA TABLOSU",
)

ws["A2"] = headers.get("colHeaderPlanetDisk", "PlanetDISK® Ünitesi")
ws["D2"] = alt_system_title
ws["A3"] = headers.get("colHeaderMotorReducer", "Motor Redüktörü")
ws["E3"] = headers.get("colHeaderBlower", "Blower")
ws["G3"] = headers.get("colHeaderPump", "Çamur Geri Devir Pompası")

ws["D2"].fill = sys_fill
for r in range(4, 12):
    ws.cell(r, 4).fill = sys_fill


# ==========================================
# 4. SATIR ETİKETLERİ (A VE D SÜTUNLARI)
# ==========================================
# PlanetDISK Tarafı Sol Etiketler (A4:A11)
ws["A4"] = headers.get("rowLabelQty", "Ünite / Ekipman Adedi")
ws["A5"] = headers.get("rowLabelPower", "Birim Motor Gücü (kW)")
ws["A6"] = headers.get("rowLabelTotalPower", "Toplam Kurulu Güç")
ws["A7"] = headers.get("rowLabelConsumptionFactor", "Anlık Güç Tüketim Oranı (%)")
ws["A8"] = headers.get("rowLabelActualPower", "Kullanılacak Gerçek Net Güç")
ws["A9"] = headers.get(
    "rowLabelPrice", f"Elektrik Birim Fiyatı ({CURRENCY_SYMBOL}/kWh)"
)
ws["A10"] = headers.get("rowLabelDailyHours", "Günlük Çalışma Süresi (saat)")
ws["A11"] = "Yıllık Çalışma Süresi" if not is_foreign else "Yearly Working Time"

# Alternatif Sistem Tarafı Orta Etiketler (D4:D11)
ws["D4"] = headers.get("rowLabelQty", "Adet")
ws["D5"] = headers.get("rowLabelPower", "Güç")
ws["D6"] = headers.get("rowLabelTotalPower", "Toplam Güç")
ws["D7"] = headers.get("rowLabelConsumptionFactor", "Tüketim Oranı (%)")
ws["D8"] = headers.get("rowLabelActualPower", "Gerçek Net Güç")
ws["D9"] = headers.get("rowLabelPrice", f"Birim Fiyat ({CURRENCY_SYMBOL}/kWh)")
ws["D10"] = headers.get("rowLabelDailyHours", "Günlük Çalışma Süresi")
ws["D11"] = "Yıllık Çalışma Süresi" if not is_foreign else "Yearly Working Time"


# ==========================================
# 5. HESAPLAMA & VERİ DOLDURMA
# ==========================================
def calc_system_metrics(raw_data, metrics_obj):
    qty = int(raw_data.get("qty", 1) or 1)
    power = float(raw_data.get("power", 0) or 0)
    factor = float(raw_data.get("consumptionFactor", 90) or 90) / 100.0

    raw_price = float(raw_data.get("price", 0.13) or 0.13)
    conv_price = raw_price * exchange_rate

    daily = int(raw_data.get("dailyHours", 24) or 24)
    yearly = int(raw_data.get("yearlyDays", 365) or 365)

    tot_pwr = float(metrics_obj.get("totalPower", qty * power) or (qty * power))
    act_pwr = float(
        metrics_obj.get("actualPower", tot_pwr * factor) or (tot_pwr * factor)
    )
    yearly_cost = float(
        metrics_obj.get("yearlyCostConverted", act_pwr * conv_price * daily * yearly)
        or (act_pwr * conv_price * daily * yearly)
    )

    return {
        "qty": qty,
        "power": power,
        "tot_pwr": tot_pwr,
        "factor": int(factor * 100),
        "act_pwr": act_pwr,
        "price": conv_price,
        "daily": daily,
        "yearly": yearly,
        "yearly_cost": yearly_cost,
    }


p_res = calc_system_metrics(planet_raw, planet_metrics)
b_res = calc_system_metrics(blower_raw, blower_metrics)
pu_res = calc_system_metrics(pump_raw, pump_metrics)

tot_alt_cost = b_res["yearly_cost"] + pu_res["yearly_cost"]
calc_yearly_saving = tot_alt_cost - p_res["yearly_cost"]
calc_10y_saving = calc_yearly_saving * 10
maint_saving = (
    float(data_obj.get("maintenanceSaving", p_res["qty"] * 494) or (p_res["qty"] * 494))
    * exchange_rate
)
calc_total_gain = calc_10y_saving + maint_saving

# PlanetDISK (B & C)
ws["B4"] = p_res["qty"]
ws["C4"] = "pieces" if is_foreign else "adet"
ws["B5"] = (
    f"{p_res['qty']} x {fmt_num(p_res['power'], 2)} kW"
    if p_res["qty"] > 1
    else f"{fmt_num(p_res['power'], 2)} kW"
)
ws["B6"] = fmt_num(p_res["tot_pwr"], 2)
ws["C6"] = "kW/hour" if is_foreign else "kW/saat"
ws["B7"] = f"{p_res['factor']}%"
ws["B8"] = fmt_num(p_res["act_pwr"], 2)
ws["C8"] = "kW"
ws["B9"] = fmt_num(p_res["price"], 2)
ws["C9"] = f" {CURRENCY_SYMBOL}/kW"
ws["B10"] = p_res["daily"]
ws["C10"] = "hour/day" if is_foreign else "saat/gün"
ws["B11"] = p_res["yearly"]
ws["C11"] = "day/year" if is_foreign else "gün/yıl"

# Blower (E & F)
ws["E4"] = b_res["qty"]
ws["F4"] = "pieces" if is_foreign else "adet"
ws["E5"] = fmt_num(b_res["power"], 2)
ws["F5"] = "kW/hour" if is_foreign else "kW/saat"
ws["E6"] = fmt_num(b_res["tot_pwr"], 2)
ws["F6"] = "kW/hour" if is_foreign else "kW/saat"
ws["E7"] = f"{b_res['factor']}%"
ws["E8"] = fmt_num(b_res["act_pwr"], 2)
ws["F8"] = "kW"
ws["E9"] = fmt_num(b_res["price"], 2)
ws["F9"] = f" {CURRENCY_SYMBOL}/kW"
ws["E10"] = b_res["daily"]
ws["F10"] = "hour/day" if is_foreign else "saat/gün"
ws["E11"] = b_res["yearly"]
ws["F11"] = "day/year" if is_foreign else "gün/yıl"

# Pump (G & H)
ws["G4"] = pu_res["qty"]
ws["H4"] = "piece" if is_foreign else "adet"
ws["G5"] = fmt_num(pu_res["power"], 2)
ws["H5"] = "kW/hour" if is_foreign else "kW/saat"
ws["G6"] = fmt_num(pu_res["tot_pwr"], 2)
ws["H6"] = "kW/hour" if is_foreign else "kW/saat"
ws["G7"] = f"{pu_res['factor']}%"
ws["G8"] = fmt_num(pu_res["act_pwr"], 2)
ws["H8"] = "kW"
ws["G9"] = fmt_num(pu_res["price"], 2)
ws["H9"] = f" {CURRENCY_SYMBOL}/kW"
ws["G10"] = pu_res["daily"]
ws["H10"] = "hour/day" if is_foreign else "saat/gün"
ws["G11"] = pu_res["yearly"]
ws["H11"] = "day/year" if is_foreign else "gün/yıl"


# ==========================================
# 6. ÖZET VE KART METİNLERİ (G12 VE H12 BİRLEŞTİRİLDİ)
# ==========================================
ws["A12"] = headers.get("rowLabelYearlyCost", "Yıllık Tüketim Maliyeti")

yearly_cost_row = rendered_content.get("yearlyCostRow", {})

ws["B12"] = yearly_cost_row.get(
    "planet", f"{fmt_num(round(p_res['yearly_cost']), 0)} {CURRENCY_SYMBOL} /yıl"
)
ws["C12"] = ""

ws["D12"] = alt_system_title + (
    " Energy Consumption" if is_foreign else " Enerji Gideri"
)

# E12:F12 birleşmiş durumda
ws["E12"] = yearly_cost_row.get(
    "blower", f"{fmt_num(round(b_res['yearly_cost']), 0)} {CURRENCY_SYMBOL} /yıl"
)

# G12:H12 alanını birleştirip ortalıyoruz
ws.merge_cells("G12:H12")
ws["G12"] = yearly_cost_row.get(
    "pump", f"{fmt_num(round(pu_res['yearly_cost']), 0)} {CURRENCY_SYMBOL} /yıl"
)
ws["G12"].alignment = Alignment(horizontal="center", vertical="center")

# Tasarruf Kartları
ws["A13"] = headers.get(
    "card1Title", f'"{alt_system_title}"\'ne Kıyasla Yıllık Enerji Tasarrufu:'
)
ws["E13"] = summary_cards.get(
    "yearlySavingFormatted",
    f"{fmt_num(round(calc_yearly_saving), 0)} {CURRENCY_SYMBOL} /yıl",
)

ws["A14"] = headers.get("card2Title", "Sistem Ömrü Boyunca 10 Yıllık Elektrik Kazancı:")
ws["E14"] = summary_cards.get(
    "tenYearsSavingFormatted",
    f"{fmt_num(round(calc_10y_saving), 0)} {CURRENCY_SYMBOL} /10 yıl",
)

ws["A15"] = headers.get(
    "card3Title",
    "Blower, Difüzör Yenileme ve Bakım Maliyetleri Dahil Yaklaşık Toplam Tasarruf (10 Yıl)",
)
round_gain = round(calc_total_gain / 1000.0) * 1000 if calc_total_gain else 0
ws["E15"] = summary_cards.get(
    "totalGainWithMaintenanceFormatted", f"{fmt_num(round_gain, 0)} {CURRENCY_SYMBOL}"
)


# ==========================================
# 7. KAYDETME
# ==========================================
wb.save(OUTPUT_FILE)

print(f"✅ Referans Alınan Şablon : {SOURCE_FILE}")
print(f"🚀 Oluşturulan Yeni Tablo : {OUTPUT_FILE}")
