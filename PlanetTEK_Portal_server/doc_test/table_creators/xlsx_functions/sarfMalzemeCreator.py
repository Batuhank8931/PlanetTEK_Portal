import json
import os
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ==========================================
# 1. DOSYA & JSON VERİ OKUMA (GÜNCELLENDİ)
# ==========================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# 💾 Kaydedilecek hedef dosya ve klasör (xlsx1)
OUTPUT_DIR = os.path.join(BASE_DIR, "..", "tables", "xlsx1")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "sarf_malzeme.xlsx")

# Target klasör (xlsx1) yoksa otomatik oluştur
os.makedirs(OUTPUT_DIR, exist_ok=True)

JSON_PATH = os.path.join(BASE_DIR, "..", "..", "formData.json")

if not os.path.exists(JSON_PATH):
    raise FileNotFoundError(f"Hata: {JSON_PATH} dosyası bulunamadı!")

with open(JSON_PATH, "r", encoding="utf-8") as f:
    form_data = json.load(f)

customer_info = form_data.get("customerInfo", {})
sarf_data_obj = form_data.get("tables", {}).get("sarfmalzemettablosu", {})

rows_data = sarf_data_obj.get("rows", [])
grand_total_json = sarf_data_obj.get("grandTotal", 0)

# --- DİL, PARA BİRİMİ VE BİRİM SİSTEMİ DİNAMİKLERİ ---
is_foreign = customer_info.get("teklifDili") == "Yabancı"
currency_code = customer_info.get("currency", "EUR").upper()
unit_system = customer_info.get("unitSystem", "metric").upper()

if currency_code == "USD":
    CURRENCY_SYMBOL = "$"
elif currency_code == "TRY":
    CURRENCY_SYMBOL = "₺"
else:
    CURRENCY_SYMBOL = "€"


def fmt_num(val, decimals=2):
    """Sayıyı teklifDili ve maksimum 'decimals' haneye göre formatlar."""
    try:
        val = float(val)
    except (ValueError, TypeError):
        val = 0.0

    formatted = f"{val:,.{decimals}f}"

    if not is_foreign:
        formatted = formatted.replace(",", "X").replace(".", ",").replace("X", ".")

    return formatted


def fix_unit_text(unit_str):
    """JSON'dan gelen birim metinlerindeki sabit '€' simgesini dinamik CURRENCY_SYMBOL ile değiştirir."""
    if not unit_str:
        return ""

    fixed = str(unit_str).replace("€", CURRENCY_SYMBOL)

    if unit_system == "US":
        fixed = fixed.replace("kg", "lbs").replace("ton", "tons").replace("lt", "gal")

    return fixed


LANG = {
    "title": (
        "CONSUMABLES and MAINTENANCE COST"
        if is_foreign
        else "SARF MALZEME VE BAKIM GİDERLERİ"
    ),
    "cols": [
        "Description" if is_foreign else "Açıklama",
        "Total Amount" if is_foreign else "Toplam Miktar",
        "Consumption" if is_foreign else "Tüketim",
        "Unit Price" if is_foreign else "Birim Fiyat",
        "Total Price" if is_foreign else "Toplam Tutar",
    ],
    "grand_total": "GRAND TOTAL" if is_foreign else "GENEL TOPLAM",
    "year_unit": (
        f"{CURRENCY_SYMBOL}/year" if is_foreign else f"{CURRENCY_SYMBOL}/yıl"
    ),
}

# ==========================================
# 2. EXCEL SIFIRDAN HAZIRLIK
# ==========================================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "SarfMalzeme"
ws.views.sheetView[0].showGridLines = True

# --- STİLLER ---
FILL_YELLOW = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
FILL_LEVEL_0 = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")
FILL_LEVEL_1 = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
FILL_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

font_main_title = Font(name="Calibri", size=9, bold=True, italic=True)
font_bold = Font(name="Calibri", size=9, bold=True)
font_regular = Font(name="Calibri", size=9, bold=False)

thin_s = Side(border_style="thin", color="000000")
double_s = Side(border_style="double", color="000000")

# ==========================================
# 3. BAŞLIK ALANI (A1:E2)
# ==========================================
ws.row_dimensions[1].height = 20
ws.merge_cells("A1:E1")
ws["A1"] = LANG["title"]
ws["A1"].font = font_main_title
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

for col in range(1, 6):
    ws.cell(row=1, column=col).fill = FILL_YELLOW
    ws.cell(row=1, column=col).border = Border(
        top=double_s,
        bottom=double_s,
        left=double_s if col == 1 else None,
        right=double_s if col == 5 else None,
    )

ws.row_dimensions[2].height = 18
for col_idx, text in enumerate(LANG["cols"], 1):
    cell = ws.cell(row=2, column=col_idx, value=text)
    cell.font = font_bold
    cell.fill = FILL_YELLOW
    align_h = "left" if col_idx == 1 else "center"
    cell.alignment = Alignment(horizontal=align_h, vertical="center")

for col in range(1, 6):
    ws.cell(row=2, column=col).border = Border(
        top=thin_s,
        bottom=double_s,
        left=double_s if col == 1 else thin_s,
        right=double_s if col == 5 else thin_s,
    )

# ==========================================
# 4. VERİ SATIRLARI DÖNGÜSÜ
# ==========================================
current_row = 3
calculated_grand_total = 0.0

for row_item in rows_data:
    label = row_item.get("label", "").strip()
    is_header = row_item.get("isHeader", False)
    is_sub_header = row_item.get("isSubHeader", False)
    is_light = row_item.get("isLight", False)

    if is_header and label.upper() in [
        "SARF MALZEME VE BAKIM GİDERLERİ",
        "CONSUMABLES AND MAINTENANCE COST",
    ]:
        continue

    ws.row_dimensions[current_row].height = 18

    if is_header or (is_sub_header and not is_light):
        ws.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=5,
        )
        cell = ws.cell(row=current_row, column=1, value=label)
        cell.font = font_bold
        for c in range(1, 6):
            c_cell = ws.cell(row=current_row, column=c)
            c_cell.fill = FILL_LEVEL_0
            c_cell.border = Border(
                left=double_s if c == 1 else None,
                right=double_s if c == 5 else None,
                top=None,
                bottom=None,
            )

    elif is_sub_header and is_light:
        ws.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=5,
        )
        cell = ws.cell(row=current_row, column=1, value=f"  {label}")
        cell.font = font_bold
        for c in range(1, 6):
            c_cell = ws.cell(row=current_row, column=c)
            c_cell.fill = FILL_LEVEL_1
            c_cell.border = Border(
                left=double_s if c == 1 else None,
                right=double_s if c == 5 else None,
                top=None,
                bottom=None,
            )

    else:
        qty = float(row_item.get("qty", 1) or 1)
        qty_unit = row_item.get("qtyUnit", "")

        consumption = float(row_item.get("consumption", 0) or 0)
        consumption_unit = row_item.get("consumptionUnit", "")

        unit_price = float(row_item.get("unitPrice", 0) or 0)
        price_unit = row_item.get("priceUnit", "€/kg")

        row_total = qty * consumption * unit_price
        calculated_grand_total += row_total

        fixed_price_unit = fix_unit_text(price_unit)
        fixed_cons_unit = fix_unit_text(consumption_unit)

        str_qty = f"{fmt_num(qty, 0) if qty.is_integer() else fmt_num(qty, 1)} {qty_unit}".strip()
        str_cons = f"{fmt_num(consumption, 2)} {fixed_cons_unit}".strip()
        str_price = f"{fmt_num(unit_price, 0) if unit_price.is_integer() else fmt_num(unit_price, 2)} {fixed_price_unit}".strip()
        str_total = f"{fmt_num(row_total, 2)} {LANG['year_unit']}"

        ws.cell(row=current_row, column=1, value=f"  {label}").alignment = Alignment(
            horizontal="left", vertical="center"
        )
        ws.cell(row=current_row, column=2, value=str_qty).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws.cell(row=current_row, column=3, value=str_cons).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws.cell(row=current_row, column=4, value=str_price).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws.cell(row=current_row, column=5, value=str_total).alignment = Alignment(
            horizontal="center", vertical="center"
        )

        for c in range(1, 6):
            c_cell = ws.cell(row=current_row, column=c)
            c_cell.fill = FILL_WHITE
            c_cell.font = font_regular

            left_b = double_s if c == 1 else thin_s
            right_b = double_s if c == 5 else thin_s

            c_cell.border = Border(left=left_b, right=right_b, top=None, bottom=None)

    current_row += 1

# ==========================================
# 5. GENEL TOPLAM SATIRI
# ==========================================
ws.row_dimensions[current_row].height = 20
ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)

lbl_grand = ws.cell(row=current_row, column=1, value=LANG["grand_total"])
lbl_grand.font = font_bold
lbl_grand.alignment = Alignment(horizontal="right", vertical="center")

final_total = grand_total_json if grand_total_json > 0 else calculated_grand_total
str_grand_total = f"{fmt_num(final_total, 2)} {LANG['year_unit']}"

val_grand = ws.cell(row=current_row, column=5, value=str_grand_total)
val_grand.font = font_bold
val_grand.alignment = Alignment(horizontal="center", vertical="center")

for c in range(1, 6):
    c_cell = ws.cell(row=current_row, column=c)
    c_cell.fill = FILL_YELLOW
    c_cell.border = Border(
        top=thin_s,
        bottom=double_s,
        left=double_s if c == 1 else None,
        right=double_s if c == 5 else None,
    )

# ==========================================
# 6. KOLON GENİŞLİKLERİ VE KAYDETME (xlsx1'e)
# ==========================================
ws.column_dimensions["A"].width = 38
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 25
ws.column_dimensions["D"].width = 18
ws.column_dimensions["E"].width = 22

wb.save(OUTPUT_FILE)
print(f"🚀 Excel Sarf Malzeme Tablosu Sıfırdan Çizilerek Kaydedildi: {OUTPUT_FILE}")
