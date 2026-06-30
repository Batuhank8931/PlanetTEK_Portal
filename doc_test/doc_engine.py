import os
import openpyxl
import re
from openpyxl.utils import get_column_letter

def apply_tint(hex_color, tint):
    """Excel'in tema renklerine uyguladığı Tint (Açıklık/Koyuluk) oranını işler."""
    if not hex_color or tint == 0.0:
        return hex_color
        
    if not re.match(r'^[0-9A-F]{6}$', hex_color):
        return None

    r = int(hex_color[0:2], 16)
    g = int(hex_color[2:4], 16)
    b = int(hex_color[4:6], 16)
    
    if tint < 0:
        r = int(r * (1.0 + tint))
        g = int(g * (1.0 + tint))
        b = int(b * (1.0 + tint))
    else:
        r = int(r + (255 - r) * tint)
        g = int(g + (255 - g) * tint)
        b = int(b + (255 - b) * tint)
        
    return f"{max(0, min(255, r)):02X}{max(0, min(255, g)):02X}{max(0, min(255, b)):02X}"

def clean_hex_color(cell, cell_fill):
    """Excel hücre dolgusunu tema rengi, tint kırılımları ve akıllı hiyerarşik fallback ile çözer."""
    if not cell_fill or not cell_fill.start_color:
        return get_hierarchy_fallback_color(cell)
        
    color_obj = cell_fill.start_color
    hex_str = None
    
    if color_obj.type == 'theme' and hasattr(color_obj, 'theme') and color_obj.theme is not None:
        theme_id = color_obj.theme
        if theme_id == 1 or theme_id == 3: 
            hex_str = "A6A6A6" if theme_id == 1 else "D9D9D9"
        elif theme_id == 0:
            hex_str = "FFFFFF"
            
    if hasattr(color_obj, 'rgb') and isinstance(color_obj.rgb, str) and color_obj.rgb != "00000000":
        t_hex = color_obj.rgb.upper()
        if len(t_hex) == 8:
            hex_str = t_hex[2:]
        elif len(t_hex) == 6:
            hex_str = t_hex

    if not hex_str or not re.match(r'^[0-9A-F]{6}$', hex_str):
        return get_hierarchy_fallback_color(cell)
        
    if hasattr(color_obj, 'tint') and isinstance(color_obj.tint, (int, float)) and color_obj.tint:
        hex_str = apply_tint(hex_str, color_obj.tint)
        
    return hex_str

def get_hierarchy_fallback_color(cell):
    """Metin hiyerarşisine göre grinin tonlarını bağlar."""
    val = str(cell.value).strip() if cell.value is not None else ""
    if not val:
        return None
        
    if re.match(r'^\d+\.\s+[A-Z\s]+', val):
        return "7F7F7F"
    elif re.match(r'^\d+\.\d+\.\s+', val):
        return "D9D9D9"
    elif re.match(r'^\d+\.\d+\.\d+\.\s+', val):
        return "F2F2F2"
        
    return None

def map_border_style(excel_border_side):
    """Excel border tiplerini jilet gibi spesifik CSS formatına çevirir."""
    if not excel_border_side or excel_border_side.style is None:
        return "1px solid #1C1C1C"
        
    style = excel_border_side.style
    color = "1C1C1C"
    
    if style == "double":
        return f"3px double #{color}"
    elif style == "medium" or style == "mediumDashDot":
        return f"2px solid #{color}"
    elif style == "thick":
        return f"3px solid #{color}"
    else:
        return f"1px solid #{color}"

def get_merge_map(ws):
    """Excel'deki birleşmiş hücre haritasını çıkarır."""
    merge_map = {}
    skip_cells = set()
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        top_left_coord = f"{get_column_letter(min_col)}{min_row}"
        merge_map[top_left_coord] = {
            "rowspan": max_row - min_row + 1 if max_row > min_row else None,
            "colspan": max_col - min_col + 1 if max_col > min_col else None
        }
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                if r == min_row and c == min_col:
                    continue
                skip_cells.add(f"{get_column_letter(c)}{r}")
    return merge_map, skip_cells

def export_existing_excel_to_html(excel_path="test.xlsx", output_html_path="table.html"):
    if not os.path.exists(excel_path):
        print(f"Hata: '{excel_path}' dosyası bulunamadı!")
        return

    print("-> test.xlsx mevcut haliyle yükleniyor...")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    merge_map, skip_cells = get_merge_map(ws)

    html = ['<html><head><meta charset="utf-8">']
    html.append('<style>')
    html.append('  table { border-collapse: collapse; width: 100%; font-family: "Arial", sans-serif; font-size: 10.5px; table-layout: fixed; line-height: 1.3; }')
    html.append('  td { padding: 5px 6px; word-wrap: break-word; vertical-align: middle; }')
    html.append('</style></head><body>')
    html.append('<table>')
    
    # 🛠️ HARDKOD KANUNU: Excel'in yanlış okuduğu kaba genişlikleri çöpe atıyoruz!
    # Sırasıyla: No (1), Description (2), Piece (3), Unit Price (4), Total Price (5), Discount Rate (6), Total After Discount (7)
    # Toplam oran tam 650px (A4'e milimetrik sığma referansı)
    target_widths = [35, 275, 40, 70, 75, 65, 90]
    
    html.append('<colgroup>')
    for w_px in target_widths:
        html.append(f'  <col style="width: {w_px}px;">')
    html.append('%s' % '</colgroup>')

    for r_idx in range(1, ws.max_row + 1):
        row_height = ws.row_dimensions[r_idx].height
        height_style = f"height: {int(row_height * 1.3)}px;" if row_height else ""
        html.append(f'  <tr style="{height_style}">')
        
        for c_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(c_idx)
            coord = f"{col_letter}{r_idx}"
            
            if coord in skip_cells:
                continue
                
            cell = ws.cell(row=r_idx, column=c_idx)
            val = str(cell.value) if cell.value is not None else ""
            
            bg_color = clean_hex_color(cell, cell.fill)
            is_bold = "bold" if (cell.font and cell.font.bold) else "normal"
            is_italic = "italic" if (cell.font and cell.font.italic) else "normal"
            
            font_color = "000000"
            if cell.font and cell.font.color and isinstance(cell.font.color.rgb, str):
                f_color = cell.font.color.rgb.upper()
                font_color = f_color[2:] if len(f_color) == 8 else f_color
                if not re.match(r'^[0-9A-F]{6}$', font_color):
                    font_color = "000000"

            font_size = f"font-size: {cell.font.size}pt;" if (cell.font and cell.font.size) else ""

            align = "left"
            if cell.alignment and cell.alignment.horizontal:
                align = cell.alignment.horizontal
                if align == "centerContinuous": align = "center"
            else:
                if c_idx in [1, 3, 6]: align = "center" # No, Piece, Discount rate ortalı
                elif c_idx in [4, 5, 7]: align = "right" # Fiyat kolonları sağa yaslı

            border_top = map_border_style(cell.border.top) if cell.border else "1px solid #1C1C1C"
            border_bottom = map_border_style(cell.border.bottom) if cell.border else "1px solid #1C1C1C"
            border_left = map_border_style(cell.border.left) if cell.border else "1px solid #1C1C1C"
            border_right = map_border_style(cell.border.right) if cell.border else "1px solid #1C1C1C"

            styles = [
                f"text-align: {align}",
                f"font-weight: {is_bold}",
                f"font-style: {is_italic}",
                f"color: #{font_color}",
                font_size,
                f"border-top: {border_top}",
                f"border-bottom: {border_bottom}",
                f"border-left: {border_left}",
                f"border-right: {border_right}"
            ]
            if bg_color:
                styles.append(f"background-color: #{bg_color}")
            style_str = "; ".join(styles)

            merge_attrs = ""
            if coord in merge_map:
                m_info = merge_map[coord]
                if m_info["rowspan"]: merge_attrs += f' rowspan="{m_info["rowspan"]}"'
                if m_info["colspan"]: merge_attrs += f' colspan="{m_info["colspan"]}"'

            html.append(f'    <td{merge_attrs} style="{style_str};">{val}</td>')
            
        html.append('  </tr>')
        
    html.append('</table></body></html>')

    with open(output_html_path, "w", encoding="utf-8") as f:
        f.write("\n".join(html))
        
    print(f"\n🎉 GENİŞLİKLER SIKIŞTIRILDI! Kolonlar daraltılarak '{output_html_path}' üretildi.")

if __name__ == "__main__":
    export_existing_excel_to_html()