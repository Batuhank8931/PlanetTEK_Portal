import os
import re
import openpyxl
from openpyxl.utils import get_column_letter


def apply_tint(hex_color, tint):
    """Excel'in tema renklerine uyguladığı Tint oranını işler."""
    if not hex_color or tint == 0.0:
        return hex_color

    if not re.match(r"^[0-9A-F]{6}$", hex_color):
        return None

    r = int(hex_color[0:2], 16)
    g = int(hex_color[2:4], 16)
    b = int(hex_color[4:6], 16)

    if tint < 0:
        r = int(r * (1.0 + tint))
        g = int(g * (1.0 + tint))
        b = int(b * (1.0 + tint))
    else:
        r = int(r + (255 - r) * tint)
        g = int(g + (255 - g) * tint)
        b = int(b + (255 - b) * tint)

    return (
        f"{max(0, min(255, r)):02X}{max(0, min(255, g)):02X}{max(0, min(255, b)):02X}"
    )


def clean_hex_color(cell, cell_fill):
    """Excel hücre dolgusunu çözümler."""
    if not cell_fill or not cell_fill.start_color:
        return None

    color_obj = cell_fill.start_color
    hex_str = None

    if (
        color_obj.type == "theme"
        and hasattr(color_obj, "theme")
        and color_obj.theme is not None
    ):
        theme_id = color_obj.theme
        if theme_id in [1, 3]:
            hex_str = "A6A6A6" if theme_id == 1 else "D9D9D9"
        elif theme_id == 0:
            hex_str = "FFFFFF"

    if (
        hasattr(color_obj, "rgb")
        and isinstance(color_obj.rgb, str)
        and color_obj.rgb != "00000000"
    ):
        t_hex = color_obj.rgb.upper()
        if len(t_hex) == 8:
            hex_str = t_hex[2:]
        elif len(t_hex) == 6:
            hex_str = t_hex

    if not hex_str or not re.match(r"^[0-9A-F]{6}$", hex_str):
        return None

    if (
        hasattr(color_obj, "tint")
        and isinstance(color_obj.tint, (int, float))
        and color_obj.tint
    ):
        hex_str = apply_tint(hex_str, color_obj.tint)

    return hex_str


def map_border_style(excel_border_side):
    """Excel border tiplerini CSS karşılığına çevirir."""
    if not excel_border_side or excel_border_side.style is None:
        return None

    style = excel_border_side.style

    # Renk bilgisi varsa al yoksa siyah varsay
    color = "000000"
    if (
        excel_border_side.color
        and hasattr(excel_border_side.color, "rgb")
        and isinstance(excel_border_side.color.rgb, str)
    ):
        c_rgb = excel_border_side.color.rgb.upper()
        color = c_rgb[2:] if len(c_rgb) == 8 else c_rgb
        if not re.match(r"^[0-9A-F]{6}$", color):
            color = "000000"

    if style == "double":
        return f"3px double #{color}"
    elif style in ["medium", "mediumDashDot"]:
        return f"2px solid #{color}"
    elif style == "thick":
        return f"3px solid #{color}"
    elif style == "thin":
        return f"1px solid #{color}"
    else:
        return f"1px solid #{color}"


def get_merge_info(ws):
    """Excel birleşmiş hücre haritasını ve sınır bilgilerini çıkarır."""
    merge_map = {}
    skip_cells = set()

    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        top_left_coord = f"{get_column_letter(min_col)}{min_row}"

        merge_map[top_left_coord] = {
            "rowspan": max_row - min_row + 1 if max_row > min_row else None,
            "colspan": max_col - min_col + 1 if max_col > min_col else None,
            "max_row": max_row,
            "max_col": max_col,
            "min_row": min_row,
            "min_col": min_col,
        }

        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                if r == min_row and c == min_col:
                    continue
                skip_cells.add(f"{get_column_letter(c)}{r}")

    return merge_map, skip_cells


def convert_xlsx_to_html(excel_path, output_html_path):
    """Kapak Excel dosyasını (kapak.xlsx) alır ve birebir HTML'e çevirir."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    merge_map, skip_cells = get_merge_info(ws)

    html = ['<!DOCTYPE html>\n<html><head><meta charset="utf-8">']
    html.append("<style>")
    html.append(
        '  table { border-collapse: collapse; width: 100%; font-family: "Calibri", "Arial", sans-serif; font-size: 9pt; table-layout: fixed; line-height: 1.2; }'
    )
    html.append(
        "  td { padding: 5px 8px; word-wrap: break-word; vertical-align: middle; }"
    )
    html.append("</style></head><body>")
    html.append("<table>")

    # Kolon Genişlikleri
    html.append("<colgroup>")
    for c_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(c_idx)
        excel_width = ws.column_dimensions[col_letter].width or 20
        px_width = int(excel_width * 8.5)
        html.append(f'  <col style="width: {px_width}px;">')
    html.append("</colgroup>")

    for r_idx in range(1, ws.max_row + 1):
        row_height = ws.row_dimensions[r_idx].height
        height_style = f"height: {int(row_height * 1.33)}px;" if row_height else ""
        html.append(f'  <tr style="{height_style}">')

        for c_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(c_idx)
            coord = f"{col_letter}{r_idx}"

            if coord in skip_cells:
                continue

            cell = ws.cell(row=r_idx, column=c_idx)
            raw_val = cell.value if cell.value is not None else ""
            val_html = str(raw_val).replace("\n", "<br>")

            bg_color = clean_hex_color(cell, cell.fill)
            is_bold = "bold" if (cell.font and cell.font.bold) else "normal"
            is_italic = "italic" if (cell.font and cell.font.italic) else "normal"

            font_color = "000000"
            if cell.font and cell.font.color and isinstance(cell.font.color.rgb, str):
                f_color = cell.font.color.rgb.upper()
                font_color = f_color[2:] if len(f_color) == 8 else f_color
                if not re.match(r"^[0-9A-F]{6}$", font_color):
                    font_color = "000000"

            font_size = (
                f"font-size: {cell.font.size}pt;"
                if (cell.font and cell.font.size)
                else "font-size: 9pt;"
            )

            align = "left"
            if cell.alignment and cell.alignment.horizontal:
                align = cell.alignment.horizontal

            styles = [
                f"text-align: {align}",
                f"font-weight: {is_bold}",
                f"font-style: {is_italic}",
                f"color: #{font_color}",
                font_size,
            ]

            # Kenarlık Çözümlemesi
            top_border = cell.border.top if cell.border else None
            bottom_border = cell.border.bottom if cell.border else None
            left_border = cell.border.left if cell.border else None
            right_border = cell.border.right if cell.border else None

            # Birleşmiş Hücrelerin Sınır Kenarlıklarını Al (A1:B1 gibi birleşmeler için)
            if coord in merge_map:
                m_info = merge_map[coord]
                bottom_cell = ws.cell(row=m_info["max_row"], column=c_idx)
                right_cell = ws.cell(row=r_idx, column=m_info["max_col"])

                if bottom_cell.border and bottom_cell.border.bottom:
                    bottom_border = bottom_cell.border.bottom
                if right_cell.border and right_cell.border.right:
                    right_border = right_cell.border.right

            t = map_border_style(top_border)
            b = map_border_style(bottom_border)
            l = map_border_style(left_border)
            r = map_border_style(right_border)

            if t:
                styles.append(f"border-top: {t}")
            if b:
                styles.append(f"border-bottom: {b}")
            if l:
                styles.append(f"border-left: {l}")
            if r:
                styles.append(f"border-right: {r}")

            if bg_color:
                styles.append(f"background-color: #{bg_color}")

            style_str = "; ".join(styles)

            merge_attrs = ""
            if coord in merge_map:
                m_info = merge_map[coord]
                if m_info["rowspan"]:
                    merge_attrs += f' rowspan="{m_info["rowspan"]}"'
                if m_info["colspan"]:
                    merge_attrs += f' colspan="{m_info["colspan"]}"'

            html.append(f'    <td{merge_attrs} style="{style_str};">{val_html}</td>')

        html.append("  </tr>")

    html.append("</table></body></html>")

    with open(output_html_path, "w", encoding="utf-8") as f:
        f.write("\n".join(html))


def process_kapak_html():
    """Yalnızca kapak.xlsx dosyasını hedef alıp kapak.html üretir."""
    base_dir = os.path.dirname(os.path.abspath(__file__))

    excel_path = os.path.join(base_dir, "..", "tables", "xlsx1", "kapak.xlsx")
    output_html_path = os.path.join(base_dir, "..", "tables", "htmls", "kapak.html")

    if not os.path.exists(excel_path):
        print(f"Hata: '{excel_path}' bulunamadı!")
        return

    try:
        convert_xlsx_to_html(excel_path, output_html_path)
        print("  ✅ KAPAK HTML Dönüşümü Başarılı: kapak.xlsx -> kapak.html")
    except Exception as e:
        print(f"  ❌ KAPAK HTML Dönüşüm Hatası: {e}")


if __name__ == "__main__":
    process_kapak_html()
