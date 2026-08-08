import os
import re
import openpyxl
from openpyxl.utils import get_column_letter


def apply_tint(hex_color, tint):
    """Excel'in tema renklerine uyguladığı Tint oranını işler."""
    if not hex_color or tint == 0.0:
        return hex_color

    if not re.match(r"^[0-9A-F]{6}$", hex_color):
        return None

    r = int(hex_color[0:2], 16)
    g = int(hex_color[2:4], 16)
    b = int(hex_color[4:6], 16)

    if tint < 0:
        r = int(r * (1.0 + tint))
        g = int(g * (1.0 + tint))
        b = int(b * (1.0 + tint))
    else:
        r = int(r + (255 - r) * tint)
        g = int(g + (255 - g) * tint)
        b = int(b + (255 - b) * tint)

    return (
        f"{max(0, min(255, r)):02X}{max(0, min(255, g)):02X}{max(0, min(255, b)):02X}"
    )


def clean_hex_color(cell, cell_fill):
    """Excel hücre dolgusunu çözümler."""
    if not cell_fill or not cell_fill.start_color:
        return get_hierarchy_fallback_color(cell)

    color_obj = cell_fill.start_color
    hex_str = None

    if (
        color_obj.type == "theme"
        and hasattr(color_obj, "theme")
        and color_obj.theme is not None
    ):
        theme_id = color_obj.theme
        if theme_id in [1, 3]:
            hex_str = "A6A6A6" if theme_id == 1 else "D9D9D9"
        elif theme_id == 0:
            hex_str = "FFFFFF"

    if (
        hasattr(color_obj, "rgb")
        and isinstance(color_obj.rgb, str)
        and color_obj.rgb != "00000000"
    ):
        t_hex = color_obj.rgb.upper()
        if len(t_hex) == 8:
            hex_str = t_hex[2:]
        elif len(t_hex) == 6:
            hex_str = t_hex

    if not hex_str or not re.match(r"^[0-9A-F]{6}$", hex_str):
        return get_hierarchy_fallback_color(cell)

    if (
        hasattr(color_obj, "tint")
        and isinstance(color_obj.tint, (int, float))
        and color_obj.tint
    ):
        hex_str = apply_tint(hex_str, color_obj.tint)

    return hex_str


def get_hierarchy_fallback_color(cell):
    """Metin hiyerarşisine göre grinin tonlarını bağlar."""
    val = str(cell.value).strip() if cell.value is not None else ""
    if not val:
        return None

    if re.match(r"^\d+\.\s+[A-Z\s]+", val):
        return "A6A6A6"
    elif re.match(r"^\d+\.\d+\.\s+", val):
        return "C0C0C0"
    elif re.match(r"^\d+\.\d+\.\d+\.\s+", val):
        return "D9D9D9"

    return None


def map_border_style(excel_border_side):
    """Excel border tiplerini CSS karşılığına çevirir."""
    if not excel_border_side or excel_border_side.style is None:
        return None

    style = excel_border_side.style
    color = "000000"

    if style == "double":
        return f"3px double #{color}"
    elif style in ["medium", "mediumDashDot"]:
        return f"2px solid #{color}"
    elif style == "thick":
        return f"3px solid #{color}"
    elif style == "thin":
        return f"1px solid #{color}"
    else:
        return f"1px solid #{color}"


def get_merge_map(ws):
    """Excel'deki birleşmiş hücre haritasını çıkarır."""
    merge_map = {}
    skip_cells = set()
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        top_left_coord = f"{get_column_letter(min_col)}{min_row}"
        merge_map[top_left_coord] = {
            "rowspan": max_row - min_row + 1 if max_row > min_row else None,
            "colspan": max_col - min_col + 1 if max_col > min_col else None,
        }
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                if r == min_row and c == min_col:
                    continue
                skip_cells.add(f"{get_column_letter(c)}{r}")
    return merge_map, skip_cells


def format_currency_value(val, number_format):
    """Değer sayısal ve format varsa birim işler."""
    if isinstance(val, (int, float)):
        formatted = f"{val:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        if "₺" in number_format:
            return f"{formatted} ₺"
        elif "$" in number_format:
            return f"$ {formatted}"
        elif "€" in number_format:
            return f"{formatted} €"
        return formatted
    return str(val)


def convert_xlsx_to_html(excel_path, output_html_path):
    """On Yıllık Maliyet Excel dosyasını alır ve birebir HTML'e çevirir."""
    wb_data = openpyxl.load_workbook(excel_path, data_only=True)
    wb_formula = openpyxl.load_workbook(excel_path, data_only=False)

    ws_data = wb_data.active
    ws_formula = wb_formula.active

    merge_map, skip_cells = get_merge_map(ws_data)

    html = ['<!DOCTYPE html>\n<html><head><meta charset="utf-8">']
    html.append("<style>")
    html.append(
        '  table { border-collapse: collapse; width: 100%; font-family: "Calibri", "Arial", sans-serif; font-size: 8pt; table-layout: fixed; line-height: 1.2; }'
    )
    html.append(
        "  td { padding: 4px 5px; word-wrap: break-word; vertical-align: middle; }"
    )
    html.append("</style></head><body>")
    html.append("<table>")

    # Kolon Genişlikleri
    html.append("<colgroup>")
    for c_idx in range(1, ws_data.max_column + 1):
        col_letter = get_column_letter(c_idx)
        excel_width = ws_data.column_dimensions[col_letter].width or 12
        px_width = int(excel_width * 8.5)
        html.append(f'  <col style="width: {px_width}px;">')
    html.append("</colgroup>")

    for r_idx in range(1, ws_data.max_row + 1):
        row_height = ws_data.row_dimensions[r_idx].height
        height_style = f"height: {int(row_height * 1.33)}px;" if row_height else ""
        html.append(f'  <tr style="{height_style}">')

        for c_idx in range(1, ws_data.max_column + 1):
            col_letter = get_column_letter(c_idx)
            coord = f"{col_letter}{r_idx}"

            if coord in skip_cells:
                continue

            cell = ws_data.cell(row=r_idx, column=c_idx)
            cell_formula = wb_formula.active.cell(row=r_idx, column=c_idx)

            raw_val = cell.value

            if raw_val is None and str(cell_formula.value).startswith("="):
                raw_val = ""

            if raw_val is not None:
                if cell.number_format and (
                    "€" in cell.number_format
                    or "$" in cell.number_format
                    or "₺" in cell.number_format
                ):
                    val = format_currency_value(raw_val, cell.number_format)
                else:
                    val = str(raw_val)
            else:
                val = ""

            val_html = val.replace("\n", "<br>")

            bg_color = clean_hex_color(cell, cell.fill)
            is_bold = "bold" if (cell.font and cell.font.bold) else "normal"
            is_italic = "italic" if (cell.font and cell.font.italic) else "normal"

            font_color = "000000"
            if cell.font and cell.font.color and isinstance(cell.font.color.rgb, str):
                f_color = cell.font.color.rgb.upper()
                font_color = f_color[2:] if len(f_color) == 8 else f_color
                if not re.match(r"^[0-9A-F]{6}$", font_color):
                    font_color = "000000"

            font_size = (
                f"font-size: {cell.font.size}pt;"
                if (cell.font and cell.font.size)
                else "font-size: 8pt;"
            )

            align = "left"
            if cell.alignment and cell.alignment.horizontal:
                align = cell.alignment.horizontal
                if align == "centerContinuous":
                    align = "center"

            styles = [
                f"text-align: {align}",
                f"font-weight: {is_bold}",
                f"font-style: {is_italic}",
                f"color: #{font_color}",
                font_size,
            ]

            if cell.border:
                t = map_border_style(cell.border.top)
                b = map_border_style(cell.border.bottom)
                l = map_border_style(cell.border.left)
                r = map_border_style(cell.border.right)

                if t:
                    styles.append(f"border-top: {t}")
                if b:
                    styles.append(f"border-bottom: {b}")
                if l:
                    styles.append(f"border-left: {l}")
                if r:
                    styles.append(f"border-right: {r}")

            if bg_color:
                styles.append(f"background-color: #{bg_color}")
            style_str = "; ".join(styles)

            merge_attrs = ""
            if coord in merge_map:
                m_info = merge_map[coord]
                if m_info["rowspan"]:
                    merge_attrs += f' rowspan="{m_info["rowspan"]}"'
                if m_info["colspan"]:
                    merge_attrs += f' colspan="{m_info["colspan"]}"'

            html.append(f'    <td{merge_attrs} style="{style_str};">{val_html}</td>')

        html.append("  </tr>")

    html.append("</table></body></html>")

    with open(output_html_path, "w", encoding="utf-8") as f:
        f.write("\n".join(html))


def process_on_yillik_maliyet_html():
    """Yalnızca on_yillik_maliyet.xlsx dosyasını hedef alıp on_yillik_maliyet.html üretir."""
    base_dir = os.path.dirname(os.path.abspath(__file__))

    if os.path.exists("on_yillik_maliyet.xlsx"):
        excel_path = os.path.abspath("on_yillik_maliyet.xlsx")
        os.makedirs(os.path.join(base_dir, "tables", "htmls"), exist_ok=True)
        output_html_path = os.path.join(
            base_dir, "tables", "htmls", "on_yillik_maliyet.html"
        )
    else:
        excel_path = os.path.join(
            base_dir, "..", "tables", "xlsx1", "on_yillik_maliyet.xlsx"
        )
        output_html_path = os.path.join(
            base_dir, "..", "tables", "htmls", "on_yillik_maliyet.html"
        )
        os.makedirs(os.path.dirname(output_html_path), exist_ok=True)

    if not os.path.exists(excel_path):
        excel_path = "on_yillik_maliyet.xlsx"
        output_html_path = "on_yillik_maliyet.html"

    if not os.path.exists(excel_path):
        print(f"Hata: '{excel_path}' bulunamadı!")
        return

    try:
        convert_xlsx_to_html(excel_path, output_html_path)
        print(
            f"  ✅ On Yillik Maliyet HTML Dönüşümü Başarılı: {os.path.basename(excel_path)} -> {os.path.basename(output_html_path)}"
        )
    except Exception as e:
        print(f"  ❌ On Yillik Maliyet HTML Dönüşüm Hatası: {e}")


if __name__ == "__main__":
    process_on_yillik_maliyet_html()
