import os
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 1. Klasör ve Dosya Yollarının Yönetimi (Mac & Ubuntu Uyumlu)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TARGET_FILE = os.path.join(BASE_DIR, "tables", "capex.xlsx")
JSON_PATH = os.path.join(BASE_DIR, "..", "formData.json")

# --- formData.json'dan Veriyi Okuma ---
if not os.path.exists(JSON_PATH):
    raise FileNotFoundError(f"Hata: {JSON_PATH} dosyası bulunamadı!")

with open(JSON_PATH, "r", encoding="utf-8") as f:
    form_data = json.load(f)

data = form_data.get("tables", {}).get("capextablosu", {}).get("rows", [])

if not data:
    print("Uyarı: 'tables.capextablosu.rows' altında veri bulunamadı veya liste boş!")

# --- Excel Hazırlığı (Sıfırdan Temiz Sayfa Oluşturma) ---
if not os.path.exists(TARGET_FILE):
    print(f"Hata: {TARGET_FILE} bulunamadı! Yeni bir tane oluşturuluyor...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Capex"
    os.makedirs(os.path.dirname(TARGET_FILE), exist_ok=True)
else:
    wb = openpyxl.load_workbook(TARGET_FILE)
    if "Capex" in wb.sheetnames:
        del wb["Capex"]
    ws = wb.create_sheet(title="Capex")

# Gridlines (Kılavuz çizgileri) açık olsun
ws.views.sheetView[0].showGridLines = True

# --- Stil ve Renk Tanımlamaları ---
FILL_HEADER_YELLOW = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

FILL_LEVEL_0 = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")
FILL_LEVEL_1 = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
FILL_LEVEL_2 = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
FILL_WHITE   = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

FONT_BASE = {"name": "Calibri", "size": 8}
font_level_0 = Font(bold=True, italic=True, **FONT_BASE)
font_level_1 = Font(bold=True, **FONT_BASE)
font_level_2 = Font(bold=False, italic=True, **FONT_BASE)
font_item    = Font(bold=False, **FONT_BASE)
font_item_bold_italic = Font(bold=True, italic=True, **FONT_BASE) # Optional / Supply Locally İçin
font_header_title = Font(name="Calibri", size=9, bold=True, italic=True)

# Kenarlık Tanımları
thin_side = Side(border_style="thin", color="000000")
double_side = Side(border_style="double", color="000000")
border_thin_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
border_empty = Border() # Kenarlıksız

EURO_FORMAT = '#,##0.00" €"'
PERCENT_FORMAT = '0%'

# ==========================================
# 2. ÜST BİLGİ / BAŞLIK ALANI (Satır 1, 2, 3)
# ==========================================

ws.row_dimensions[1].height = 18
ws.row_dimensions[2].height = 18
ws.row_dimensions[3].height = 18

for r in range(1, 4):
    for c in range(1, 8):
        cell = ws.cell(row=r, column=c)
        cell.fill = FILL_HEADER_YELLOW
        cell.font = font_header_title
        
        t_side = thin_side if r == 1 else None
        b_side = thin_side if r == 3 else None
        l_side = thin_side if c == 1 else None
        r_side = thin_side if c == 7 else None
        
        cell.border = Border(top=t_side, bottom=b_side, left=l_side, right=r_side)

# --- 1. Satır: CAPEX (A1:G1) ---
ws.merge_cells("A1:G1")
ws["A1"] = "CAPEX"
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

# --- 2. Satır: İSKİ (A2:E2) & Offer Number (F2:G2) ---
ws.merge_cells("A2:E2")
ws["A2"] = "İSKİ"
ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells("F2:G2")
ws["F2"] = "Offer Number: 2026 / 3500"
ws["F2"].alignment = Alignment(horizontal="center", vertical="center")

# --- 3. Satır: A ve B Birleştirilecek (A3:B3) ---
ws.merge_cells("A3:B3")
ws["A3"] = "Offer Reference Number : YDD R0 01 01 2026 8 MX 1 70 25 0"
ws["A3"].alignment = Alignment(horizontal="center", vertical="center")

# ==========================================
# 3. KOLON BAŞLIKLARI (Satır 4)
# ==========================================
headers = [
    "No", 
    "Description", 
    "Piece", 
    "Unit Price", 
    "Total Price", 
    "Discount rate", 
    "Total Price after Discount"
]

ws.row_dimensions[4].height = 25

for col_idx, header_text in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col_idx, value=header_text)
    cell.font = Font(name="Calibri", size=8, bold=True)
    cell.fill = FILL_HEADER_YELLOW
    cell.border = border_thin_all
    
    if col_idx == 2:
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    else:
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ==========================================
# 4. VERİ DÖNGÜSÜ (Satır 5'ten Başlayarak)
# ==========================================
current_row = 5

for item in data:
    is_header = (item.get("piece") == 0 and item.get("unitPrice") == 0) or item.get("type") in [0, 1, 2]
    label_value = item.get("label", "")
    
    if item.get("type") == 0:
        label_value = label_value.upper()

    ws.cell(row=current_row, column=1, value=item.get("computedNo", ""))
    ws.cell(row=current_row, column=2, value=label_value)
    
    line_count = str(label_value).count('\n') + 1
    calculated_height = max(18, line_count * 12.5)
    ws.row_dimensions[current_row].height = calculated_height

    if is_header:
        if item.get("type") == 0 or "." not in item.get("computedNo", "").strip("."):
            row_fill = FILL_LEVEL_0
            row_font = font_level_0
        elif item.get("computedNo", "").count(".") == 2:
            row_fill = FILL_LEVEL_1
            row_font = font_level_1
        else:
            row_fill = FILL_LEVEL_2
            row_font = font_level_2
            
        for col in range(1, 8):
            cell = ws.cell(row=current_row, column=col)
            cell.fill = row_fill
            cell.font = row_font
            cell.border = border_empty
            cell.alignment = Alignment(horizontal="center" if col == 1 else "left", vertical="center", wrap_text=True)
            
    else:
        if item.get("isUrgent"):
            for col in range(1, 8):
                cell = ws.cell(row=current_row, column=col)
                cell.fill = FILL_WHITE
                cell.font = font_item
                cell.border = border_empty
                if col == 1:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col == 2:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        else:
            ws.cell(row=current_row, column=3, value=item.get("piece"))
            
            raw_discount = item.get("discount") if item.get("discount") is not None else item.get("discountRate", 0)
            try:
                discount_val = float(raw_discount)
            except (ValueError, TypeError):
                discount_val = 0.0

            discount_rate = discount_val / 100.0 if discount_val > 1 else discount_val

            ws.cell(row=current_row, column=4, value=item.get("unitPrice"))
            ws.cell(row=current_row, column=6, value=discount_rate)
            
            # --- isOptional veya isLocalSupply Kontrolü ---
            is_special_text = item.get("isOptional") or item.get("isLocalSupply")
            text_value = item.get("netTotal") if is_special_text else ""

            if is_special_text:
                # Total Price (Kolon 5) ve Total Price after Discount (Kolon 7) hücrelerine netTotal basılıyor
                ws.cell(row=current_row, column=5, value=text_value)
                ws.cell(row=current_row, column=7, value=text_value)
            else:
                ws.cell(row=current_row, column=5, value=f"=C{current_row}*D{current_row}")
                ws.cell(row=current_row, column=7, value=f"=E{current_row}*(1-F{current_row})")
                
            for col in range(1, 8):
                cell = ws.cell(row=current_row, column=col)
                cell.fill = FILL_WHITE
                cell.border = border_empty
                
                # Biçimlendirme & Font Ayarları
                if col in [5, 7] and is_special_text:
                    cell.font = font_item_bold_italic  # Bold ve Italic yapılıyor
                else:
                    cell.font = font_item
                
                if col == 1:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col == 2:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                elif col == 3:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.number_format = '#,##0'
                elif col == 4:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.number_format = EURO_FORMAT
                elif col == 5:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if not is_special_text:
                        cell.number_format = EURO_FORMAT
                elif col == 6:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.number_format = PERCENT_FORMAT
                elif col == 7:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if not is_special_text:
                        cell.number_format = EURO_FORMAT

    current_row += 1

# ==========================================
# 5. URGENT GRUPLARINI BİRLEŞTİRME (C:G)
# ==========================================
search_row = 5
urgent_groups = []
current_group = []

for item in data:
    if item.get("isUrgent"):
        current_group.append((search_row, item))
    else:
        if current_group:
            urgent_groups.append(current_group)
            current_group = []
    search_row += 1
if current_group:
    urgent_groups.append(current_group)

for group in urgent_groups:
    start_r = group[0][0]
    end_r = group[-1][0]
    unit_data_text = group[0][1].get("unitData", "")
    
    ws.merge_cells(start_row=start_r, start_column=3, end_row=end_r, end_column=7)
    merged_cell = ws.cell(row=start_r, column=3, value=unit_data_text)
    merged_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    merged_cell.font = font_item

# ==========================================
# 6. EN DİŞ ÇERÇEVEYİ ÇİFT ÇİZGİ (DOUBLE) YAPMA
# ==========================================
max_row = current_row - 1
max_col = 7

for r in range(1, max_row + 1):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=r, column=c)
        
        top = cell.border.top.style if cell.border and cell.border.top else None
        bottom = cell.border.bottom.style if cell.border and cell.border.bottom else None
        left = cell.border.left.style if cell.border and cell.border.left else None
        right = cell.border.right.style if cell.border and cell.border.right else None
        
        if r == 1:
            top = 'double'
        if r == max_row:
            bottom = 'double'
        if c == 1:
            left = 'double'
        if c == max_col:
            right = 'double'
            
        cell.border = Border(
            top=Side(border_style=top, color="000000") if top else None,
            bottom=Side(border_style=bottom, color="000000") if bottom else None,
            left=Side(border_style=left, color="000000") if left else None,
            right=Side(border_style=right, color="000000") if right else None
        )

# 7. Dinamik Sütun Genişlikleri
for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.row > 3 and cell.value:
            lines = str(cell.value).split('\n')
            for line in lines:
                if len(line) > max_len:
                    max_len = len(line)
    
    if col_letter == "B":
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 20), 50)
    else:
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

wb.save(TARGET_FILE)
print(f"Special text kuralları (Bold & Italic) uygulandı ve Excel kaydedildi: {TARGET_FILE}")