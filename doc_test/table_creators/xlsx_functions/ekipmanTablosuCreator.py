import json
import os
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ==========================================
# 1. DOSYA & JSON VERİ OKUMA (GÜNCELLENDİ)
# ==========================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# 💾 Kaydedilecek hedef dosya ve klasör (xlsx1)
OUTPUT_DIR = os.path.join(BASE_DIR, "..", "tables", "xlsx1")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "ekipman_tablosu.xlsx")

# Target klasör (xlsx1) yoksa otomatik oluştur
os.makedirs(OUTPUT_DIR, exist_ok=True)

JSON_PATH = os.path.join(BASE_DIR, "..", "..", "formData.json")

if not os.path.exists(JSON_PATH):
    raise FileNotFoundError(f"Hata: {JSON_PATH} dosyası bulunamadı!")

with open(JSON_PATH, "r", encoding="utf-8") as f:
    form_data = json.load(f)

customer_info = form_data.get("customerInfo", {})
ekipman_data = form_data.get("tables", {}).get("ekipantablosu", {})

if not ekipman_data:
    ekipman_data = form_data.get("tables", {}).get("ekipmantablosu", {})

content_list = ekipman_data.get("content", [])

# ==========================================
# 2. SIFIRDAN ÇALIŞMA KİTABI / SAYFASI OLUŞTURMA
# ==========================================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Ekipman Tablosu"

# Kılavuz çizgilerini görünür yap
ws.views.sheetView[0].showGridLines = True

# ==========================================
# 3. RENK VE STİL TANIMLARI
# ==========================================
font_main_header = Font(name="Calibri", size=11, bold=True, color="000000")
font_equip_header = Font(name="Calibri", size=11, bold=True, color="000000")
font_spec_label = Font(name="Calibri", size=10, bold=False, italic=True, color="000000")
font_spec_value = Font(
    name="Calibri", size=10, bold=False, italic=False, color="000000"
)

fill_main = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")
fill_equip = PatternFill(start_color="D6D4CA", end_color="D6D4CA", fill_type="solid")
fill_spec_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
fill_spec_even = PatternFill(
    start_color="D8D8D8", end_color="D8D8D8", fill_type="solid"
)

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

side_thin = Side(style="thin", color="000000")

# ==========================================
# 4. TABLO VERİLERİNİ DİNAMİK İŞLEME
# ==========================================
blocks = []
current_block = None

for item in content_list:
    item_type = item.get("type", "")
    if item_type in ["main", "equip"]:
        if current_block:
            blocks.append(current_block)
        current_block = {"header": item, "specs": []}
    elif item_type == "spec":
        if current_block:
            current_block["specs"].append(item)

if current_block:
    blocks.append(current_block)

current_row = 1

for b_idx, block in enumerate(blocks):
    header_item = block["header"]
    specs = block["specs"]
    h_type = header_item.get("type", "")
    h_label = header_item.get("label", "")

    if current_row > 1:
        current_row += 1

    total_specs = len(specs)

    # 1. BAŞLIK SATIRI YAZMA
    ws.merge_cells(
        start_row=current_row,
        start_column=1,
        end_row=current_row,
        end_column=2,
    )
    cell_a = ws.cell(current_row, 1)
    cell_b = ws.cell(current_row, 2)

    if h_type == "main":
        cell_a.value = str(h_label).upper()
        cell_a.font = font_main_header
        cell_a.alignment = align_center
        cell_a.fill = fill_main
        cell_b.fill = fill_main
        ws.row_dimensions[current_row].height = 24
    else:
        cell_a.value = str(h_label)
        cell_a.font = font_equip_header
        cell_a.alignment = align_left
        cell_a.fill = fill_equip
        cell_b.fill = fill_equip
        ws.row_dimensions[current_row].height = 20

    has_no_specs = total_specs == 0

    cell_a.border = Border(
        left=side_thin,
        top=side_thin,
        bottom=side_thin if has_no_specs else None,
    )
    cell_b.border = Border(
        right=side_thin,
        top=side_thin,
        bottom=side_thin if has_no_specs else None,
    )

    current_row += 1

    # 2. SPEC SATIRLARINI YAZMA
    for s_idx, spec in enumerate(specs):
        is_last_spec = s_idx == total_specs - 1
        s_label = spec.get("label", "")
        s_value = spec.get("value", "")

        c_a = ws.cell(current_row, 1, value=str(s_label))
        c_b = ws.cell(current_row, 2, value=str(s_value))

        row_fill = fill_spec_even if s_idx % 2 == 1 else fill_spec_odd

        c_a.font = font_spec_label
        c_a.alignment = align_left
        c_a.fill = row_fill

        c_b.font = font_spec_value
        c_b.alignment = align_left
        c_b.fill = row_fill

        c_a.border = Border(left=side_thin, bottom=side_thin if is_last_spec else None)
        c_b.border = Border(right=side_thin, bottom=side_thin if is_last_spec else None)

        val_len = len(str(s_value))
        if val_len > 80:
            ws.row_dimensions[current_row].height = 36
        elif val_len > 50:
            ws.row_dimensions[current_row].height = 26
        else:
            ws.row_dimensions[current_row].height = 18

        current_row += 1

# ==========================================
# 5. SÜTUN GENİŞLİKLERİ VE KAYDETME (xlsx1'e)
# ==========================================
ws.column_dimensions["A"].width = 38
ws.column_dimensions["B"].width = 40

wb.save(OUTPUT_FILE)
print(f"🚀 Sıfırdan Çizilen Ekipman Tablosu Oluşturuldu: {OUTPUT_FILE}")
