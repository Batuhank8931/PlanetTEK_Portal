import json
import os
import openpyxl
from openpyxl.styles import PatternFill

# ==========================================
# 1. DOSYA & JSON VERİ OKUMA (GÜNCELLENDİ)
# ==========================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# 📂 Okunacak referans/şablon dosya (xlsx0)
SOURCE_FILE = os.path.join(
    BASE_DIR, "..", "tables", "xlsx0", "enerji_karsilastirma.xlsx"
)

# 💾 Kaydedilecek hedef dosya ve klasör (xlsx1)
OUTPUT_DIR = os.path.join(BASE_DIR, "..", "tables", "xlsx1")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "enerji_karsilastirma.xlsx")

# Target klasör (xlsx1) yoksa otomatik oluştur
os.makedirs(OUTPUT_DIR, exist_ok=True)

JSON_PATH = os.path.join(BASE_DIR, "..", "..", "formData.json")

if not os.path.exists(JSON_PATH):
    raise FileNotFoundError(f"Hata: {JSON_PATH} dosyası bulunamadı!")

with open(JSON_PATH, "r", encoding="utf-8") as f:
    form_data = json.load(f)

customer_info = form_data.get("customerInfo", {})
cmp_data = form_data.get("tables", {}).get("enerjikarsilastirmatablosu", {})

if not cmp_data:
    cmp_data = form_data.get("tables", {}).get("enerjikarsilastirma", {})

data_obj = cmp_data.get("data", {})
planet_raw = data_obj.get("planet", {})
blower_raw = data_obj.get("blower", {})
pump_raw = data_obj.get("pump", {})

planet_metrics = cmp_data.get("planetMetrics", {})
blower_metrics = cmp_data.get("blowerMetrics", {})
pump_metrics = cmp_data.get("pumpMetrics", {})
rendered_summary = cmp_data.get("renderedSummary", {})

selected_system = cmp_data.get("selectedSystem", "aktif_camur")
is_mbbr = str(selected_system).lower() == "mbbr"

# --- DİL VE PARA BİRİMİ DİNAMİKLERİ ---
is_foreign = customer_info.get("teklifDili") == "Yabancı"
currency_code = customer_info.get("currency", "EUR").upper()
exchange_rate = float(customer_info.get("exchangeRate", 1.0) or 1.0)

if currency_code == "USD":
    CURRENCY_SYMBOL = "$"
elif currency_code == "TRY":
    CURRENCY_SYMBOL = "₺"
else:
    CURRENCY_SYMBOL = "€"


def fmt_num(val, decimals=2):
    """Sayıyı teklifDili'ne göre formatlar."""
    if val is None or val == "":
        return "0"
    try:
        if isinstance(val, str):
            val = val.replace(",", "")
        val_float = float(val)
    except (ValueError, TypeError):
        return str(val)

    formatted = f"{val_float:,.{decimals}f}"

    if not is_foreign:
        formatted = formatted.replace(",", "X").replace(".", ",").replace("X", ".")

    return formatted


def fix_summary_text(summary_str):
    """renderedSummary alanındaki metinleri Türkçe teklif diline uyarlar."""
    if not summary_str:
        return ""
    text = str(summary_str)
    if not is_foreign:
        text = text.replace("/year", "/yıl").replace("/10 year", "/10 yıl")
    return text


# ==========================================
# 2. ŞABLON SAYFASINI YÜKLEME (xlsx0'dan)
# ==========================================
if not os.path.exists(SOURCE_FILE):
    raise FileNotFoundError(f"Hata: Şablon dosya bulunamadı! Yol: {SOURCE_FILE}")

wb = openpyxl.load_workbook(SOURCE_FILE)

ws = wb.active
ws.views.sheetView[0].showGridLines = True

# --- DİNAMİK DOLGU RENKLERİ ---
FILL_GREEN_HEADER = PatternFill(
    start_color="92D05E", end_color="92D05E", fill_type="solid"
)
FILL_BLUE_HEADER = PatternFill(
    start_color="8EA9DB", end_color="8EA9DB", fill_type="solid"
)
sys_fill = FILL_BLUE_HEADER if is_mbbr else FILL_GREEN_HEADER

# ==========================================
# 3. BAŞLIK VE DİL METİNLERİNİ GÜNCELLEME
# ==========================================
if is_foreign:
    alt_system_title = "MBBR SYSTEM" if is_mbbr else "ACTIVATED SLUDGE SYSTEM"
    ws["A1"] = f"ENERGY COMPARISON of PlanetDISK® RBC vs {alt_system_title}"
    ws["A2"] = "PlanetDISK® Unit"
    ws["D2"] = "MBBR System" if is_mbbr else "Activated Sludge"
    ws["A3"] = "Motor Reduction Gear"
    ws["E3"] = "Blower"
    ws["G3"] = "Sludge Feed Pump"
else:
    alt_system_title = "MBBR SİSTEMİ" if is_mbbr else "AKTİF ÇAMUR SİSTEMİ"
    ws["A1"] = (
        f"PlanetDISK® DBD TEKNOLOJİSİ İLE {alt_system_title} ENERJİ KARŞILAŞTIRMA TABLOSU"
    )
    ws["A2"] = "PlanetDISK® Ünitesi"
    ws["D2"] = "MBBR Sistemi" if is_mbbr else "Klasik Aktif Çamur Sistemi"
    ws["A3"] = "Motor Redüktörü"
    ws["E3"] = "Blower"
    ws["G3"] = "Çamur Geri Devir Pompası"

ws["D2"].fill = sys_fill
for r in range(4, 12):
    ws.cell(r, 4).fill = sys_fill


# ==========================================
# 4. HESAPLAMA & PARAMETRE HESAPLAMALARI
# ==========================================
def calc_system_metrics(raw_data, metrics_obj):
    qty = int(raw_data.get("qty", 1) or 1)
    power = float(raw_data.get("power", 0) or 0)
    factor = float(raw_data.get("consumptionFactor", 90) or 90) / 100.0

    raw_price = float(raw_data.get("price", 0.13) or 0.13)
    conv_price = raw_price * exchange_rate

    daily = int(raw_data.get("dailyHours", 24) or 24)
    yearly = int(raw_data.get("yearlyDays", 365) or 365)

    tot_pwr = float(metrics_obj.get("totalPower", qty * power) or (qty * power))
    act_pwr = float(
        metrics_obj.get("actualPower", tot_pwr * factor) or (tot_pwr * factor)
    )

    yearly_cost = float(
        metrics_obj.get("yearlyCostConverted", act_pwr * conv_price * daily * yearly)
        or (act_pwr * conv_price * daily * yearly)
    )

    return {
        "qty": qty,
        "power": power,
        "tot_pwr": tot_pwr,
        "factor": int(factor * 100),
        "act_pwr": act_pwr,
        "price": conv_price,
        "daily": daily,
        "yearly": yearly,
        "yearly_cost": yearly_cost,
    }


p_res = calc_system_metrics(planet_raw, planet_metrics)
b_res = calc_system_metrics(blower_raw, blower_metrics)
pu_res = calc_system_metrics(pump_raw, pump_metrics)

# Tasarruf Hesapları
tot_alt_cost = b_res["yearly_cost"] + pu_res["yearly_cost"]
calc_yearly_saving = tot_alt_cost - p_res["yearly_cost"]
calc_10y_saving = calc_yearly_saving * 10
maint_saving = (
    float(data_obj.get("maintenanceSaving", p_res["qty"] * 494) or (p_res["qty"] * 494))
    * exchange_rate
)
calc_total_gain = calc_10y_saving + maint_saving

# ==========================================
# 5. TEKNİK PARAMETRE HÜCRELERİNİ DOLDURMA
# ==========================================
# PlanetDISK (B & C Kolonları)
ws["B4"] = p_res["qty"]
ws["C4"] = "pieces" if is_foreign else "adet"
ws["B5"] = (
    f"{p_res['qty']} x {fmt_num(p_res['power'], 2)} kW"
    if p_res["qty"] > 1
    else f"{fmt_num(p_res['power'], 2)} kW"
)
ws["B6"] = fmt_num(p_res["tot_pwr"], 2)
ws["C6"] = "kW/hour" if is_foreign else "kW/saat"
ws["B7"] = f"{p_res['factor']}%"
ws["B8"] = fmt_num(p_res["act_pwr"], 2)
ws["C8"] = "kW"
ws["B9"] = fmt_num(p_res["price"], 2)
ws["C9"] = f" {CURRENCY_SYMBOL}/kW"
ws["B10"] = p_res["daily"]
ws["C10"] = "hour/day" if is_foreign else "saat/gün"
ws["B11"] = p_res["yearly"]
ws["C11"] = "day/year" if is_foreign else "gün/yıl"

# Blower (E & F Kolonları)
ws["E4"] = b_res["qty"]
ws["F4"] = "pieces" if is_foreign else "adet"
ws["E5"] = fmt_num(b_res["power"], 2)
ws["F5"] = "kW/hour" if is_foreign else "kW/saat"
ws["E6"] = fmt_num(b_res["tot_pwr"], 2)
ws["F6"] = "kW/hour" if is_foreign else "kW/saat"
ws["E7"] = f"{b_res['factor']}%"
ws["E8"] = fmt_num(b_res["act_pwr"], 2)
ws["F8"] = "kW"
ws["E9"] = fmt_num(b_res["price"], 2)
ws["F9"] = f" {CURRENCY_SYMBOL}/kW"
ws["E10"] = b_res["daily"]
ws["F10"] = "hour/day" if is_foreign else "saat/gün"
ws["E11"] = b_res["yearly"]
ws["F11"] = "day/year" if is_foreign else "gün/yıl"

# Pump (G & H Kolonları)
ws["G4"] = pu_res["qty"]
ws["H4"] = "piece" if is_foreign else "adet"
ws["G5"] = fmt_num(pu_res["power"], 2)
ws["H5"] = "kW/hour" if is_foreign else "kW/saat"
ws["G6"] = fmt_num(pu_res["tot_pwr"], 2)
ws["H6"] = "kW/hour" if is_foreign else "kW/saat"
ws["G7"] = f"{pu_res['factor']}%"
ws["G8"] = fmt_num(pu_res["act_pwr"], 2)
ws["H8"] = "kW"
ws["G9"] = fmt_num(pu_res["price"], 2)
ws["H9"] = f" {CURRENCY_SYMBOL}/kW"
ws["G10"] = pu_res["daily"]
ws["H10"] = "hour/day" if is_foreign else "saat/gün"
ws["G11"] = pu_res["yearly"]
ws["H11"] = "day/year" if is_foreign else "gün/yıl"

# ==========================================
# 6. ÖZET VE RENDERED SUMMARY METİNLERİ (Satır 12 - 15)
# ==========================================
ws["A12"] = (
    "PlanetDISK® Unit Energy Consumption"
    if is_foreign
    else "PlanetDISK® Ünitesi Enerji Gideri"
)
ws["B12"] = rendered_summary.get(
    "planetYearlyCostFormatted",
    f"{fmt_num(round(p_res['yearly_cost']), 0)} {CURRENCY_SYMBOL}",
)
ws["C12"] = "/year" if is_foreign else "/yıl"

ws["D12"] = (
    ("MBBR Energy Consumption" if is_mbbr else "Activated Sludge Energy Consumption")
    if is_foreign
    else ("MBBR Enerji Gideri" if is_mbbr else "Klasik Aktif Çamur Enerji Gideri")
)
ws["E12"] = rendered_summary.get(
    "blowerYearlyCostFormatted",
    f"{fmt_num(round(b_res['yearly_cost']), 0)} {CURRENCY_SYMBOL}",
)
ws["F12"] = "/year" if is_foreign else "/yıl"

ws["G12"] = rendered_summary.get(
    "pumpYearlyCostFormatted",
    f"{fmt_num(round(pu_res['yearly_cost']), 0)} {CURRENCY_SYMBOL}",
)
ws["H12"] = "/year" if is_foreign else "/yıl"

ws["A13"] = (
    "Comparing PlanetDISK® Unit with other systems yearly, electric power saving is"
    if is_foreign
    else "PlanetDISK® diğer atıksu arıtma sistemleri ile karşılaştırıldığında, enerji tasarrufu"
)
ws["E13"] = fix_summary_text(
    rendered_summary.get(
        "yearlySavingFormatted",
        f"{fmt_num(round(calc_yearly_saving), 0)} {CURRENCY_SYMBOL} /year",
    )
)

ws["A14"] = (
    "For 10 years electric power saving price is equal to "
    if is_foreign
    else "Tasarruf tutarı 10 yılda "
)
ws["E14"] = fix_summary_text(
    rendered_summary.get(
        "tenYearsSavingFormatted",
        f"{fmt_num(round(calc_10y_saving), 0)} {CURRENCY_SYMBOL} /10 year",
    )
)

ws["A15"] = (
    "and together with blower and diffusers maintenance cost, it will be approx"
    if is_foreign
    else "Blower ve Difüzör Bakım Masraflarıyla birlikte bu rakam"
)

round_gain = round(calc_total_gain / 1000.0) * 1000 if calc_total_gain else 0
ws["E15"] = fix_summary_text(
    rendered_summary.get(
        "totalGainWithMaintenanceFormatted",
        f"{fmt_num(round_gain, 0)} {CURRENCY_SYMBOL}",
    )
)

# ==========================================
# 7. KAYDETME (xlsx1'e)
# ==========================================
wb.save(OUTPUT_FILE)

print(f"✅ Referans Alınan Şablon : {SOURCE_FILE}")
print(f"🚀 Oluşturulan Yeni Tablo : {OUTPUT_FILE}")
