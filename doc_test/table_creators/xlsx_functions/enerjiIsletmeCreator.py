import json
import os
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ==========================================
# 1. DOSYA & JSON VERİ OKUMA (GÜNCELLENDİ)
# ==========================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# 💾 Kaydedilecek hedef dosya ve klasör (xlsx1)
OUTPUT_DIR = os.path.join(BASE_DIR, "..", "tables", "xlsx1")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "enerji_isletme.xlsx")

# Target klasör (xlsx1) yoksa otomatik oluştur
os.makedirs(OUTPUT_DIR, exist_ok=True)

JSON_PATH = os.path.join(BASE_DIR, "..", "..", "formData.json")

if not os.path.exists(JSON_PATH):
    raise FileNotFoundError(f"Hata: {JSON_PATH} dosyası bulunamadı!")

with open(JSON_PATH, "r", encoding="utf-8") as f:
    form_data = json.load(f)

customer_info = form_data.get("customerInfo", {})
enerji_data_obj = form_data.get("tables", {}).get("enerjiisletmettablosu", {})

rows_data = enerji_data_obj.get("rows", [])
summary_data = enerji_data_obj.get("summary", {})

# Üst Alan Verileri
raw_energy_price = enerji_data_obj.get("inputEnergyPrice")
raw_hydraulic = enerji_data_obj.get("inputHydraulic")

# --- DİL, PARA BİRİMİ VE BİRİM SİSTEMİ DİNAMİKLERİ ---
is_foreign = customer_info.get("teklifDili") == "Yabancı"
currency_code = customer_info.get("currency", "EUR").upper()
unit_system = customer_info.get("unitSystem", "metric").upper()

if currency_code == "USD":
    CURRENCY_SYMBOL = "$"
    CENT_SYMBOL = "$ cent"
elif currency_code == "TRY":
    CURRENCY_SYMBOL = "₺"
    CENT_SYMBOL = "Kuruş"
else:
    CURRENCY_SYMBOL = "€"
    CENT_SYMBOL = "€ cent"

if unit_system == "US":
    FLOW_UNIT = "GPH"
    VOL_UNIT = "gal"
    KWH_PER_VOL_UNIT = "kW.hour / gal" if is_foreign else "kW.saat / gal"
else:
    FLOW_UNIT = "m³/hour" if is_foreign else "m³/saat"
    VOL_UNIT = "m³"
    KWH_PER_VOL_UNIT = "kW.hour / m³" if is_foreign else "kW.saat / m³"


def fmt_num(val, decimals=2):
    if val is None or val == "":
        return ""
    try:
        if isinstance(val, str):
            val = val.replace(",", "")
        val_float = float(val)
    except (ValueError, TypeError):
        return str(val)

    formatted = f"{val_float:,.{decimals}f}"

    if not is_foreign:
        formatted = formatted.replace(",", "X").replace(".", ",").replace("X", ".")

    return formatted


LANG = {
    "title": ("ENERGY OPERATION COST" if is_foreign else "ENERJİ İŞLETME GİDERLERİ"),
    "total_flow_label": (
        "Total Hydraulic Load" if is_foreign else "Toplam Hidrolik Yük"
    ),
    "energy_price_label": (
        "Energy Price for 1 kW.hour" if is_foreign else "1 kW.saat Enerji Bedeli"
    ),
    "main_header": ("MECHANICAL EQUIPMENTS" if is_foreign else "MEKANİK EKİPMANLAR"),
    "cols": {
        "unit": "Unit" if is_foreign else "Adet",
        "p_inst": ("Unit Installed\nPower" if is_foreign else "Birim Kurulu\nGüç"),
        "p_tot": ("Total Installed\nPower" if is_foreign else "Toplam Kurulu\nGüç"),
        "p_cons": ("Power\nConsumed" if is_foreign else "Güç Kullanım\nYüzdesi"),
        "working": "Daily Working " if is_foreign else "Günlük Çalışma",
        "elec": ("Electricity\nConsumption" if is_foreign else "Elektrik\nTüketimi"),
        "p_unit": "kW",
        "cons_unit": "%",
        "work_unit": "hour" if is_foreign else "saat",
        "elec_unit": "kW.hour/day" if is_foreign else "kW.saat/gün",
    },
    "tot_elec": (
        "TOTAL ELECTRICITY CONSUMPTION" if is_foreign else "TOPLAM ELEKTRİK TÜKETİMİ"
    ),
    "per_m3_kw": (
        f"Electricity consumption per 1 {VOL_UNIT} wastewater"
        if is_foreign
        else f"1 {VOL_UNIT} atıksu başına elektrik tüketimi"
    ),
    "per_m3_cost": (
        f"Electricity consumption per 1 {VOL_UNIT} wastewater"
        if is_foreign
        else f"1 {VOL_UNIT} atıksu başına elektrik giderleri"
    ),
    "cost_label": ("Electricity Consumption Cost" if is_foreign else "Elektrik Gideri"),
    "day_unit": (
        f"{CURRENCY_SYMBOL} / day" if is_foreign else f"{CURRENCY_SYMBOL} / gün"
    ),
    "year_unit": (
        f"{CURRENCY_SYMBOL} / year" if is_foreign else f"{CURRENCY_SYMBOL} / yıl"
    ),
    "footer": (
        f'*** "Energy Price for kw/hour = 13 {CENT_SYMBOL}" is the rate in TURKEY & given for comparison purposes'
        if is_foreign
        else f'*** "1 kW/saat = 13 {CENT_SYMBOL}" TÜRKİYE şartları baz alınmış olup karşılaştırma amaçlıdır.'
    ),
}

design_flow_val = (
    raw_hydraulic
    if raw_hydraulic is not None
    else customer_info.get("dizaynDebisi", 2.92)
)
energy_price_val = (
    raw_energy_price
    if raw_energy_price is not None
    else customer_info.get("elektrikBirimFiyatCent", 13)
)

# ==========================================
# 2. EXCEL SAYFASI HAZIRLIĞI (SIFIR SAYFA)
# ==========================================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "EnerjiIsletme"
ws.views.sheetView[0].showGridLines = True

# --- STİLLER ---
FILL_YELLOW = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
FILL_LEVEL_0 = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")
FILL_LEVEL_1 = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
FILL_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

font_main_title = Font(name="Calibri", size=9, bold=True, italic=True)
font_bold = Font(name="Calibri", size=9, bold=True)
font_regular = Font(name="Calibri", size=9, bold=False)
font_footer = Font(name="Calibri", size=9, bold=True, italic=True, color="FF0000")

thin_s = Side(border_style="thin", color="000000")
double_s = Side(border_style="double", color="000000")

# ==========================================
# 3. ÜST SARI BAŞLIK ALANI (A1:H4)
# ==========================================
ws.row_dimensions[1].height = 20
ws.merge_cells("A1:H1")
ws["A1"] = LANG["title"]
ws["A1"].font = font_main_title
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

for col in range(1, 9):
    ws.cell(row=1, column=col).fill = FILL_YELLOW
    ws.cell(row=1, column=col).border = Border(
        top=double_s,
        bottom=double_s,
        left=double_s if col == 1 else None,
        right=double_s if col == 8 else None,
    )

ws.row_dimensions[2].height = 18
ws.row_dimensions[3].height = 18

ws["A2"] = LANG["total_flow_label"]
ws["A2"].font = font_bold
ws["A2"].alignment = Alignment(horizontal="right", vertical="center")

ws.merge_cells("B2:C2")
ws["B2"] = f"{fmt_num(design_flow_val, 2)} {FLOW_UNIT}"
ws["B2"].font = font_regular
ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

ws["A3"] = LANG["energy_price_label"]
ws["A3"].font = font_bold
ws["A3"].alignment = Alignment(horizontal="right", vertical="center")

ws.merge_cells("B3:C3")
ws["B3"] = f"{fmt_num(energy_price_val, 2)} {CENT_SYMBOL}"
ws["B3"].font = font_regular
ws["B3"].alignment = Alignment(horizontal="center", vertical="center")

for r in [2, 3]:
    for c in range(1, 4):
        ws.cell(row=r, column=c).border = Border(
            top=thin_s,
            bottom=thin_s,
            left=double_s if c == 1 else thin_s,
            right=thin_s,
        )

headers_map = [
    ("D2:D3", LANG["cols"]["p_inst"]),
    ("E2:E3", LANG["cols"]["p_tot"]),
    ("F2:F3", LANG["cols"]["p_cons"]),
    ("G2:G3", LANG["cols"]["working"]),
    ("H2:H3", LANG["cols"]["elec"]),
]

for rng, text in headers_map:
    ws.merge_cells(rng)
    first_cell = ws[rng.split(":")[0]]
    first_cell.value = text
    first_cell.font = font_bold
    first_cell.fill = FILL_YELLOW
    first_cell.alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )

for r in [2, 3]:
    for c in range(4, 9):
        ws.cell(row=r, column=c).fill = FILL_YELLOW
        ws.cell(row=r, column=c).border = Border(
            top=thin_s,
            bottom=thin_s,
            left=thin_s,
            right=double_s if c == 8 else thin_s,
        )

ws.row_dimensions[4].height = 18
ws.merge_cells("A4:B4")
ws["A4"] = LANG["main_header"]
ws["A4"].font = font_bold
ws["A4"].alignment = Alignment(horizontal="left", vertical="center")

sub_units = [
    (3, LANG["cols"]["unit"]),
    (4, LANG["cols"]["p_unit"]),
    (5, LANG["cols"]["p_unit"]),
    (6, LANG["cols"]["cons_unit"]),
    (7, LANG["cols"]["work_unit"]),
    (8, LANG["cols"]["elec_unit"]),
]

for col_idx, u_text in sub_units:
    cell = ws.cell(row=4, column=col_idx, value=u_text)
    cell.font = font_bold
    cell.alignment = Alignment(horizontal="center", vertical="center")

for c in range(1, 9):
    ws.cell(row=4, column=c).fill = FILL_YELLOW
    ws.cell(row=4, column=c).border = Border(
        top=thin_s,
        bottom=double_s,
        left=double_s if c == 1 else thin_s,
        right=double_s if c == 8 else thin_s,
    )

# ==========================================
# 4. VERİ SATIRLARI
# ==========================================
current_row = 5

for row_item in rows_data:
    label = str(row_item.get("label", "")).strip()
    is_header = row_item.get("isHeader", False)
    is_sub_header = row_item.get("isSubHeader", False)
    is_light = row_item.get("isLight", False)

    if is_header and label.upper() in [
        "MEKANİK EKİPMANLAR",
        "MECHANICAL EQUIPMENTS",
    ]:
        continue

    ws.row_dimensions[current_row].height = 18

    if is_header or (is_sub_header and not is_light):
        ws.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=8,
        )
        cell = ws.cell(row=current_row, column=1, value=label)
        cell.font = font_bold
        for c in range(1, 9):
            c_cell = ws.cell(row=current_row, column=c)
            c_cell.fill = FILL_LEVEL_0
            c_cell.border = Border(
                left=double_s if c == 1 else None,
                right=double_s if c == 8 else None,
                top=None,
                bottom=None,
            )

    elif is_sub_header and is_light:
        ws.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=8,
        )
        cell = ws.cell(row=current_row, column=1, value=f"  {label}")
        cell.font = font_bold
        for c in range(1, 9):
            c_cell = ws.cell(row=current_row, column=c)
            c_cell.fill = FILL_LEVEL_1
            c_cell.border = Border(
                left=double_s if c == 1 else None,
                right=double_s if c == 8 else None,
                top=None,
                bottom=None,
            )

    else:
        qty = float(row_item.get("qty", 1) or 1)
        try:
            power = float(row_item.get("power", 0) or 0)
        except (ValueError, TypeError):
            power = 0.0

        consumed = float(row_item.get("consumed", 90) or 90)
        hours = float(row_item.get("hours", 24) or 24)

        total_power = qty * power
        daily_kwh = qty * power * (consumed / 100.0) * hours

        ws.cell(row=current_row, column=1, value=f"  {label}").alignment = Alignment(
            horizontal="left", vertical="center"
        )
        ws.cell(
            row=current_row,
            column=3,
            value=fmt_num(qty, 0) if qty.is_integer() else fmt_num(qty, 1),
        ).alignment = Alignment(horizontal="center", vertical="center")

        ws.cell(
            row=current_row, column=4, value=fmt_num(power, 2)
        ).alignment = Alignment(horizontal="center", vertical="center")

        ws.cell(
            row=current_row, column=5, value=fmt_num(total_power, 2)
        ).alignment = Alignment(horizontal="center", vertical="center")

        ws.cell(
            row=current_row,
            column=6,
            value=f"{int(consumed)}%" if consumed > 1 else f"{consumed:.0%}",
        ).alignment = Alignment(horizontal="center", vertical="center")

        ws.cell(
            row=current_row,
            column=7,
            value=fmt_num(hours, 0) if hours.is_integer() else fmt_num(hours, 1),
        ).alignment = Alignment(horizontal="center", vertical="center")

        ws.cell(
            row=current_row, column=8, value=fmt_num(daily_kwh, 2)
        ).alignment = Alignment(horizontal="center", vertical="center")

        for c in range(1, 9):
            c_cell = ws.cell(row=current_row, column=c)
            c_cell.fill = FILL_WHITE
            c_cell.font = font_regular

            left_b = (
                double_s if c == 1 else (thin_s if c in [3, 4, 5, 6, 7, 8] else None)
            )
            right_b = (
                double_s if c == 8 else (thin_s if c in [2, 3, 4, 5, 6, 7] else None)
            )

            c_cell.border = Border(left=left_b, right=right_b, top=None, bottom=None)

    current_row += 1

# ==========================================
# 5. ALT ÖZET & MALİYET HESAPLARI
# ==========================================
tot_elec_val = summary_data.get("totalElectricityConsumption", {}).get("value", 0)
kwh_per_vol_val = summary_data.get("consumptionPerUnitWastewater", {}).get("value", 0)
cent_per_vol_val = summary_data.get("costPerUnitWastewater", {}).get("value", 0)
daily_cost_val = summary_data.get("dailyElectricityCost", {}).get("value", 0)
yearly_cost_val = summary_data.get("yearlyElectricityCost", {}).get("value", 0)

ws.row_dimensions[current_row].height = 20
ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
lbl_tot = ws.cell(row=current_row, column=1, value=LANG["tot_elec"])
lbl_tot.font = font_bold
lbl_tot.alignment = Alignment(horizontal="right", vertical="center")

val_tot = ws.cell(row=current_row, column=8, value=fmt_num(tot_elec_val, 2))
val_tot.font = font_bold
val_tot.alignment = Alignment(horizontal="center", vertical="center")

for c in range(1, 9):
    ws.cell(row=current_row, column=c).border = Border(
        top=thin_s,
        bottom=double_s,
        left=double_s if c == 1 else None,
        right=double_s if c == 8 else None,
    )

current_row += 1

ws.row_dimensions[current_row].height = 18
ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
ws.cell(row=current_row, column=1, value=LANG["per_m3_kw"]).font = font_bold
ws.cell(row=current_row, column=1).alignment = Alignment(
    horizontal="right", vertical="center"
)

ws.cell(row=current_row, column=7, value=fmt_num(kwh_per_vol_val, 2)).font = font_bold
ws.cell(row=current_row, column=7).alignment = Alignment(
    horizontal="center", vertical="center"
)
ws.cell(row=current_row, column=8, value=KWH_PER_VOL_UNIT).font = font_regular

for c in range(1, 9):
    ws.cell(row=current_row, column=c).border = Border(
        left=double_s if c == 1 else None, right=double_s if c == 8 else None
    )

current_row += 1

ws.row_dimensions[current_row].height = 18
ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
ws.cell(row=current_row, column=1, value=LANG["per_m3_cost"]).font = font_bold
ws.cell(row=current_row, column=1).alignment = Alignment(
    horizontal="right", vertical="center"
)

ws.cell(row=current_row, column=7, value=fmt_num(cent_per_vol_val, 2)).font = font_bold
ws.cell(row=current_row, column=7).alignment = Alignment(
    horizontal="center", vertical="center"
)
ws.cell(row=current_row, column=8, value=f"cent / {VOL_UNIT}").font = font_regular

for c in range(1, 9):
    ws.cell(row=current_row, column=c).border = Border(
        left=double_s if c == 1 else None, right=double_s if c == 8 else None
    )

current_row += 1

ws.row_dimensions[current_row].height = 36
ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
lbl_cost = ws.cell(row=current_row, column=1, value=LANG["cost_label"])
lbl_cost.font = font_bold
lbl_cost.alignment = Alignment(horizontal="right", vertical="center")

ws.cell(
    row=current_row,
    column=7,
    value=f"{fmt_num(daily_cost_val, 2)}\n{fmt_num(yearly_cost_val, 0)}",
).font = font_bold
ws.cell(row=current_row, column=7).alignment = Alignment(
    horizontal="center", vertical="center", wrap_text=True
)

ws.cell(
    row=current_row,
    column=8,
    value=f"{LANG['day_unit']}\n{LANG['year_unit']}",
).font = font_regular
ws.cell(row=current_row, column=8).alignment = Alignment(
    horizontal="left", vertical="center", wrap_text=True
)

for c in range(1, 9):
    c_cell = ws.cell(row=current_row, column=c)
    c_cell.fill = FILL_LEVEL_1
    c_cell.border = Border(
        top=thin_s,
        bottom=double_s,
        left=double_s if c == 1 else None,
        right=double_s if c == 8 else None,
    )

current_row += 1

ws.row_dimensions[current_row].height = 20
ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
footer_cell = ws.cell(row=current_row, column=1, value=LANG["footer"])
footer_cell.font = font_footer
footer_cell.alignment = Alignment(horizontal="center", vertical="center")

# ==========================================
# 6. KOLON GENİŞLİKLERİ VE KAYDETME (xlsx1'e)
# ==========================================
ws.column_dimensions["A"].width = 38
ws.column_dimensions["B"].width = 12
ws.column_dimensions["C"].width = 8
ws.column_dimensions["D"].width = 14
ws.column_dimensions["E"].width = 14
ws.column_dimensions["F"].width = 14
ws.column_dimensions["G"].width = 14
ws.column_dimensions["H"].width = 18

wb.save(OUTPUT_FILE)
print(f"🚀 Enerji İşletme tablosu sıfırdan çizilerek kaydedildi: {OUTPUT_FILE}")
